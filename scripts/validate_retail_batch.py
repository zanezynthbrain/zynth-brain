from pathlib import Path
from zipfile import ZipFile
from xml.etree import ElementTree as ET
import json, hashlib
from openpyxl import load_workbook
ROOT=Path('/home/ubuntu/zynth-brain'); OUT=ROOT/'daily_proposals'/'2026-08-21_Retail_Bilingual'; PROP=OUT/'proposals'; ASSET=OUT/'assets'
files=sorted(PROP.glob('*.docx')); assets=sorted(ASSET.glob('*.png'))
results={'document_count':len(files),'asset_count':len(assets),'english_count':len([p for p in files if p.name.endswith('_English.docx')]),'hybrid_count':len([p for p in files if 'Hybrid_English_Myanmar' in p.name]),'missing_sections':[],'myanmar_hybrid_hits':0,'bad_docx':[]}
for p in files:
    try:
        with ZipFile(p) as z:
            xml=z.read('word/document.xml')
            text=''.join(ET.fromstring(xml).itertext())
        required=['Executive decision','Insight and behaviour change','Budget packages and commercial logic','ROI scenarios and break-even logic','KPI scorecard','Risk, compliance, and cultural care','Approval ask','Source log']
        miss=[x for x in required if x not in text]
        if miss: results['missing_sections'].append({'file':p.name,'missing':miss})
        if 'မြန်မာ' in text or 'အတည်ပြု' in text: results['myanmar_hybrid_hits']+=1
    except Exception as e: results['bad_docx'].append({'file':p.name,'error':str(e)})
manifest=json.loads((OUT/'source_manifest.json').read_text(encoding='utf-8')); results['manifest_counts']={k:manifest.get(k) for k in ['proposalCount','documentCount','videoConceptCount']}; results['manifest_proposals']=len(manifest.get('proposals',[]))
wb=load_workbook(OUT/'2026-08-21_ZYNTH_Retail_Monitoring_and_Proposal_Index.xlsx',read_only=True); results['workbook_sheets']=wb.sheetnames
reg=json.loads((ROOT/'backend/outputs/variety_registry.json').read_text(encoding='utf-8')); results['registry_last_cycle']=reg['cycles'][-1]
results['sha256_manifest']=hashlib.sha256((OUT/'source_manifest.json').read_bytes()).hexdigest()
results['pass']=results['document_count']==20 and results['asset_count']==10 and results['english_count']==10 and results['hybrid_count']==10 and results['myanmar_hybrid_hits']>=10 and not results['missing_sections'] and not results['bad_docx'] and results['manifest_proposals']==10 and set(['Proposal Index','Source Log','Budget Summary']).issubset(set(results['workbook_sheets']))
(OUT/'validation_report.json').write_text(json.dumps(results,ensure_ascii=False,indent=2),encoding='utf-8'); print(json.dumps(results,ensure_ascii=False))

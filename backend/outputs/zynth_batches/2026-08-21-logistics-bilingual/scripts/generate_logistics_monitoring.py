from pathlib import Path
import json, hashlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation

ROOT=Path('/home/ubuntu/zynth-brain/backend/outputs/zynth_batches/2026-08-21-logistics-bilingual')
MON=ROOT/'monitoring'; MON.mkdir(exist_ok=True)
CAM=json.loads((ROOT/'data/campaigns.json').read_text(encoding='utf-8'))['campaigns']
COM=json.loads((ROOT/'data/commercials.json').read_text(encoding='utf-8'))['commercials']
NAVY='173A5E'; TEAL='1A6A68'; GOLD='BE8834'; PALE='EAF1F6'; WHITE='FFFFFF'; GREY='6B7280'; INK='1F2933'
fill_h=PatternFill('solid',fgColor=NAVY); fill_l=PatternFill('solid',fgColor=PALE); thin=Side(style='thin',color='D9E2F0')

def style(ws):
    ws.sheet_view.showGridLines=False; ws.column_dimensions['A'].width=3
    for row in ws.iter_rows():
        for c in row: c.alignment=Alignment(vertical='center',wrap_text=True); c.font=Font(name='Calibri',size=9,color=INK)
def title(ws,t,sub,last):
    ws.merge_cells(start_row=2,start_column=2,end_row=2,end_column=last); ws.cell(2,2,t).font=Font(name='Georgia',size=20,bold=True,color=NAVY); ws.row_dimensions[2].height=30
    ws.merge_cells(start_row=3,start_column=2,end_row=3,end_column=last); ws.cell(3,2,sub).font=Font(name='Calibri',size=9,italic=True,color=GREY); ws.row_dimensions[3].height=27
def grid(ws,row,headers,rows,widths):
    for i,h in enumerate(headers,2):
        c=ws.cell(row,i,h); c.fill=fill_h; c.font=Font(name='Calibri',size=9,bold=True,color=WHITE); c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); c.border=Border(left=thin,right=thin,top=thin,bottom=thin)
    ws.row_dimensions[row].height=28
    for rnum,record in enumerate(rows,row+1):
        for i,v in enumerate(record,2):
            c=ws.cell(rnum,i,v); c.border=Border(left=thin,right=thin,top=thin,bottom=thin); c.alignment=Alignment(vertical='center',wrap_text=True)
            if (rnum-row)%2: c.fill=PatternFill('solid',fgColor='F9FBFD')
        ws.row_dimensions[rnum].height=42
    for i,w in enumerate(widths,2): ws.column_dimensions[chr(64+i)].width=w
    return row+1,row+len(rows)
def mmk(x): return f'MMK {x/1e6:.1f}m'
def first_num(s):
    import re
    m=re.search(r'(\d+)',s); return int(m.group(1)) if m else None

wb=Workbook(); ov=wb.active; ov.title='Overview'; style(ov); title(ov,'ZYNTH Logistics Two-Hour Batch — Command View','Batch: ZYNTH-20260821-LOGISTICS-BILINGUAL | All figures are non-tax planning assumptions, not quotes, forecasts or guarantees.',9)
metrics=[('Campaign proposals','=COUNTA(Campaigns!B7:B16)','Ten distinct campaign / activation records'),('Commercial concepts','=COUNTA(Commercials!B7:B16)','Ten separate commercial/storyboard records'),('Storyboard frames','=SUM(Commercials!L7:L16)','Twelve detailed frames per commercial'),('Public-source records','=COUNTA(Research!B7:B12)','Use limitations retained'),('Recommended portfolio budget','=SUM(Campaigns!M7:M16)','MMK non-tax planning envelope'),('Pending human gates','=COUNTIF(Decision_Gates!I7:I36,"Pending")','No external action authorised')]
for ix,(label,formula,note) in enumerate(metrics):
    r=7+(ix//3)*4; c=2+(ix%3)*3
    ov.merge_cells(start_row=r,start_column=c,end_row=r,end_column=c+1); ov.cell(r,c,label).font=Font(name='Georgia',size=10,bold=True,color=NAVY); ov.cell(r,c).fill=fill_l
    ov.merge_cells(start_row=r+1,start_column=c,end_row=r+1,end_column=c+1); ov.cell(r+1,c,formula).font=Font(name='Georgia',size=17,bold=True,color=TEAL)
    if 'budget' in label.lower(): ov.cell(r+1,c).number_format='MMK #,##0'
    ov.merge_cells(start_row=r+2,start_column=c,end_row=r+2,end_column=c+1); ov.cell(r+2,c,note).font=Font(name='Calibri',size=8,italic=True,color=GREY)
    for rr in range(r,r+3):
        for cc in range(c,c+2): ov.cell(rr,cc).border=Border(left=thin,right=thin,top=thin,bottom=thin)
ov['B20']='CONTROLLED USE RULES'; ov['B20'].font=Font(name='Georgia',size=13,bold=True,color=NAVY)
for r,txt in enumerate(['• Update the same CMP/COM IDs and version fields; do not create a new tracker for revisions.','• No public-source record authorises a route, port, terminal, carrier, warehouse, capacity, price, service level, customer or partnership claim.','• No live activation, procurement, shooting, customer data capture or operational site activity proceeds without client and named human gate clearance.','• Replace planning scenarios with client-validated funnel, margin, capacity, inventory and attribution data before commercial sign-off.'],21):
    ov.merge_cells(start_row=r,start_column=2,end_row=r,end_column=9); ov.cell(r,2,txt).font=Font(name='Calibri',size=9,italic=True,color=GREY); ov.row_dimensions[r].height=23
for col,w in {'B':26,'C':16,'D':3,'E':26,'F':16,'G':3,'H':26,'I':16}.items(): ov.column_dimensions[col].width=w

ws=wb.create_sheet('Campaigns'); style(ws); title(ws,'Campaign / Activation Register','One row per campaign. Planning-only values require client replacement and approval.',22)
h=['Campaign ID','Name','Myanmar title','Format','Commercial tension','Audience / behaviour change','Conversion mechanism','Creative territory','Seasonal logic','Status','Lean MMK','Recommended MMK','Flagship MMK','Base meetings','Primary KPI','Approvals','Source IDs','Latest artifact','Current version','Notes']
rows=[]
for c in CAM:
    rows.append([c['id'],c['shortName'],c['titleMm'],c['format'],c['commercialTension'],c['audience'],c['conversionMechanism'],c['creativeTerritory'],c['seasonalLogic'],'Concept ready',c['budgetMMK']['lean'],c['budgetMMK']['recommended'],c['budgetMMK']['flagship'],first_num(c['planningOutcomes']['base'].split('/')[-1]),'; '.join(c['primaryKpis']),'; '.join(c['requiredApprovals']),', '.join(c['sourceUse']),'Drive upload pending','v1.0','Planning only'])
s,e=grid(ws,6,h,rows,[18,24,28,28,35,32,35,30,30,16,14,16,14,16,30,36,16,25,14,18]); ws.auto_filter.ref=f'B6:U{e}'; ws.freeze_panes='B7'
for col in ['L','M','N']:
    for r in range(s,e+1): ws[f'{col}{r}'].number_format='MMK #,##0'

ws=wb.create_sheet('Commercials'); style(ws); title(ws,'Standalone Commercial & Storyboard Register','COM records remain separate from their linked campaign, including production and rights decisions.',18)
h=['Commercial ID','Title','Myanmar title','Format','Linked campaign','Tension','Objective','Creative territory','Hook','CTA','Storyboard status','Frames','Status','Production gate','Latest artifact','Current version']
rows=[]
for c in COM: rows.append([c['id'],c['titleEn'],c['titleMm'],c['format'],c['linkedCampaign'],c['tension'],c['objective'],c['territory'],c['hook'],c['cta'],c['storyboardStatus'],len(c['storyboard']),'Concept storyboard ready','Client factual service/product, rights, location/site safety, privacy/security, cast, music, subtitles and end card required','Drive upload pending','v1.0'])
s,e=grid(ws,6,h,rows,[19,26,27,24,20,34,31,28,33,30,34,12,20,44,25,14]); ws.auto_filter.ref=f'B6:Q{e}'; ws.freeze_panes='B7'

ws=wb.create_sheet('Research'); style(ws); title(ws,'Research & Evidence Register','Public sources are context only. Read the use limitation before applying a source to client copy.',10)
research=[
('LOG-S01','Myanmar Logistics Institute / MIFFA','Local logistics-training website','Capability context: communication, critical thinking, customer service, organisational and problem-solving skills','No endorsement, partnership, trainer, course, venue or contact permission','https://myanmarlogisticsinstitute.com/','21 Aug 2026','Context only','Verified'),
('LOG-S02','EuroCham Myanmar','Current event page','Historic online logistics-briefing format and resilience-topic context','Member-exclusive historic event only; no route/service/partner claim','https://eurocham-myanmar.org/events/logistics-briefing-current-global-trade-transport-insights-for-businesses-in-myanmar/','12 Mar 2026','Event format','Verified'),
('LOG-S03','Myanma Port Authority','Official annual trade overview','2025 vessel-call and trade tables are published by port category','Historical data only; no current capacity, route, lead-time or port-condition claim','https://www.mpa.gov.mm/annual-trade-overview/','2025 data','Official context','Verified'),
('LOG-S04','Myanma Port Authority','Official project page','Port-planning, national logistics/transport-plan coordination and stakeholder-format context','Historic/suspended project details; not a current project status, permit or partnership','https://www.mpa.gov.mm/development_projects/project-for-formulation-of-port-master-plan/','21 Aug 2026','Planning context','Verified'),
('LOG-S05','Global New Light of Myanmar','Search-discovered local-publisher result','Association reconstitution discovery context','Page blocked by automated verification; do not use client-facing until human validation','https://www.gnlm.com.mm/miffa-reconstituted-as-myanmar-logistics-association-mla/','10 Aug 2026','Discovery only','Unverified'),
('LOG-S06','Myanmar Government','Official holiday calendar','Q4 2026 timing guardrail','Calendar only; not an operations/demand/availability forecast','https://myanmar.gov.mm/upcoming-holidays','21 Aug 2026','Calendar','Verified')]
grid(ws,6,['Source ID','Source','Type','Finding retained','Use limitation','URL','Date','Use','Status'],research,[12,28,22,36,44,52,14,20,16]); ws.auto_filter.ref='B6:J12'; ws.freeze_panes='B7'
for r in range(7,13): ws.cell(r,7).hyperlink=ws.cell(r,7).value; ws.cell(r,7).font=Font(color='0563C1',underline='single',size=9)

ws=wb.create_sheet('Decision_Gates'); style(ws); title(ws,'Human Decision Gates','Named human owners must clear each gate before external action.',10)
rows=[]
for c in CAM:
    num=c['id'].split('-')[-1]; rows += [[f'GATE-{num}-G1',c['id'],'Campaign','G1 Research Cleared','Are product/service, data, claim and evidence boundaries clear?','Client evidence owner','Before development','Pending','Approved factual brief'],[f'GATE-{num}-G4',c['id'],'Campaign','G4 Feasibility','Are site/route safety, privacy, rights, staffing, permits and escalation owners named?','Client operations owner','Before production','Pending','Approved preflight register']]
for c in COM:
    num=c['id'].split('-')[-1]; rows.append([f'GATE-{num}-G5',c['id'],'Commercial','G5 Go / No-go','Are script, service facts, cast, site, rights, system mock-ups, safety, subtitles and CTA approved?','Client brand/legal owner','Before shoot','Pending','Approved production pack'])
grid(ws,6,['Gate record','Record ID','Track','Gate','Decision question','Owner','Due / timing','Status','Evidence required'],rows,[18,20,16,18,43,25,18,16,38]); ws.auto_filter.ref='B6:J36'; ws.freeze_panes='B7'; dv=DataValidation(type='list',formula1='"Pending,Approved,Hold,Rejected,Completed"'); ws.add_data_validation(dv); dv.add('I7:I36')

ws=wb.create_sheet('Portfolio_Input'); style(ws); title(ws,'Live Tracker Import Rows','Paste after human review into the canonical ZYNTH Master Tracker.',14)
rows=[]
for c in CAM: rows.append([c['id'],'Campaign',c['shortName'],c['format'],'G2 Strategy Selected','Concept ready',c['budgetMMK']['recommended'],c['conversionMechanism'],', '.join(c['sourceUse']),'v1.0','ZYNTH-20260821-LOGISTICS-BILINGUAL','Planning/public context'])
for c in COM: rows.append([c['id'],'Commercial',c['titleEn'],c['format'],'G3 Creative Selected','Concept storyboard ready','n/a',c['cta'],'12-frame storyboard in batch package','v1.0','ZYNTH-20260821-LOGISTICS-BILINGUAL','Planning/public context'])
grid(ws,6,['Record ID','Track','Title','Format','Gate','Status','Recommended budget / n.a.','Primary conversion or CTA','Source/asset note','Version','Batch code','Classification'],rows,[20,15,29,28,20,22,20,42,28,14,32,26]); ws.auto_filter.ref='B6:M26'; ws.freeze_panes='B7'
for r in range(7,27):
    if isinstance(ws.cell(r,8).value,(int,float)): ws.cell(r,8).number_format='MMK #,##0'
chart=BarChart(); chart.type='bar'; chart.style=10; chart.title='Recommended Campaign Planning Envelope (MMK)'; chart.y_axis.title='Campaign'; chart.x_axis.title='MMK'; data=Reference(wb['Campaigns'],min_col=13,max_col=13,min_row=6,max_row=16); cats=Reference(wb['Campaigns'],min_col=3,max_col=3,min_row=7,max_row=16); chart.add_data(data,titles_from_data=True); chart.set_categories(cats); chart.height=8; chart.width=18; chart.series[0].graphicalProperties.solidFill=GOLD; ov.add_chart(chart,'B27')
for sheet in wb.worksheets: style(sheet)
out=MON/'ZYNTH-20260821-LOGISTICS-Monitoring.xlsx'; wb.save(out)
lean=sum(c['budgetMMK']['lean'] for c in CAM); rec=sum(c['budgetMMK']['recommended'] for c in CAM); flag=sum(c['budgetMMK']['flagship'] for c in CAM)
report=['# ZYNTH Logistics Monitoring Report','', '**Batch:** `ZYNTH-20260821-LOGISTICS-BILINGUAL`  ','', '## Portfolio Monitor','', '- Campaign proposals: **10**','- Standalone commercial/storyboard concepts: **10**','- Detailed storyboard frames: **120**','- Research-source records: **6**','', '## Portfolio Planning Envelopes','',f'- Lean: **{mmk(lean)}**',f'- Recommended: **{mmk(rec)}**',f'- Flagship: **{mmk(flag)}**','', '> Planning envelopes are non-tax assumptions, not supplier quotes or financial forecasts. Replace all outcomes with client data before sign-off.','', '## Monitoring Coverage','', 'The workbook tracks campaign and commercial IDs separately, conversion/CTA, budgets, required approvals, research limits, human gates, latest artifact and version.','', '## Controls','', '- Do not represent a public source as a partnership, route, port, service, venue, carrier, performance or endorsement claim.','- Route all site/route safety, confidentiality, privacy, rights, permits, data, product/service facts and operational decisions to named client owners.','- The AI Council supports structured critique; human decisions remain authoritative.']
(MON/'monitoring_report.md').write_text('\n'.join(report)+'\n',encoding='utf-8')
files=[]
for p in sorted(ROOT.rglob('*')):
    if p.is_file() and p.name!='source_manifest.json': files.append({'path':str(p.relative_to(ROOT)),'bytes':p.stat().st_size,'sha256':hashlib.sha256(p.read_bytes()).hexdigest()})
(MON/'source_manifest.json').write_text(json.dumps({'batchCode':'ZYNTH-20260821-LOGISTICS-BILINGUAL','counts':{'campaignProposals':10,'commercialConcepts':10,'storyboardFrames':120,'wordDocuments':20,'markdownDocuments':21,'researchSources':6},'files':files},ensure_ascii=False,indent=2)+'\n',encoding='utf-8')
print(out); print('recommended portfolio envelope',rec)

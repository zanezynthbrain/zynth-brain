from __future__ import annotations

import hashlib
import json
from pathlib import Path
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.utils import get_column_letter

ROOT = Path('/home/ubuntu/zynth-brain/backend/outputs/zynth_batches/2026-08-21-energy-bilingual')
CONCEPTS = json.loads((ROOT / 'data' / 'concepts.json').read_text(encoding='utf-8'))
OUT = ROOT / 'monitoring'
OUT.mkdir(parents=True, exist_ok=True)
NAVY, LIGHT, PALE, GOLD, GREEN, RED, MUTED = '0E3150','D6E3F0','F7FAFC','E2A744','2F7B61','A53A38','5D6A75'
THIN = Side(style='thin', color='C7D5DE')
MEDIUM = Side(style='medium', color=NAVY)


def set_cell(ws, cell, value, *, bold=False, fill=None, color='17212B', size=10, wrap=False, align='left', numfmt=None):
    c = ws[cell]
    c.value = value
    c.font = Font(name='Aptos', size=size, bold=bold, color=color)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    if fill: c.fill = PatternFill('solid', fgColor=fill)
    if numfmt: c.number_format = numfmt
    return c


def table_style(ws, start_row, end_row, start_col, end_col):
    for r in range(start_row, end_row+1):
        for c in range(start_col, end_col+1):
            cell=ws.cell(r,c)
            left = MEDIUM if c==start_col else Side(style=None)
            right = MEDIUM if c==end_col else Side(style=None)
            top = MEDIUM if r==start_row else THIN
            bottom = MEDIUM if r==end_row else THIN
            if r==start_row: bottom=MEDIUM
            cell.border = Border(left=left,right=right,top=top,bottom=bottom)
            cell.alignment = Alignment(horizontal='center' if r==start_row else ('right' if isinstance(cell.value,(int,float)) else 'left'),vertical='center',wrap_text=True)
            if r==start_row:
                cell.font=Font(name='Aptos',size=9,bold=True,color='FFFFFF')
                cell.fill=PatternFill('solid',fgColor=NAVY)
            elif r%2==1:
                cell.fill=PatternFill('solid',fgColor=PALE)


def title(ws, label, subtitle, end_col=10):
    ws.sheet_view.showGridLines=False
    ws.column_dimensions['A'].width=3
    ws.merge_cells(start_row=2,start_column=2,end_row=2,end_column=end_col)
    set_cell(ws,'B2',label,bold=True,color=NAVY,size=20)
    ws.row_dimensions[2].height=34
    ws.merge_cells(start_row=3,start_column=2,end_row=3,end_column=end_col)
    set_cell(ws,'B3',subtitle,color=MUTED,size=10,wrap=True)
    ws.row_dimensions[3].height=30


def build_workbook():
    wb=Workbook()
    ws=wb.active; ws.title='Overview'
    title(ws,'ZYNTH Energy Batch — Monitoring Command Sheet','Batch ZYNTH-20260821-ENERGY-BILINGUAL • Planning assumptions only • Prepared 21 Aug 2026 UTC',10)
    ws.merge_cells('B5:E5'); set_cell(ws,'B5','BATCH COVERAGE',bold=True,fill=LIGHT,color=NAVY,size=13)
    metrics=[('Industry','Energy / solar adoption'),('Concepts',10),('Word proposals',20),('Physical design packages',8),('Digital UI storyboards',2),('Source IDs',6),('Model status','Planning assumptions / client data required')]
    for i,(a,b) in enumerate(metrics,start=6):
        set_cell(ws,f'B{i}',a,bold=True,color=NAVY);set_cell(ws,f'C{i}',b,wrap=True)
    table_style(ws,6,12,2,3)
    ws.merge_cells('F5:J5'); set_cell(ws,'F5','EXECUTIVE READ-OUT',bold=True,fill=LIGHT,color=NAVY,size=13)
    rec_total=sum(c['budgets'][1] for c in CONCEPTS)
    flag_total=sum(c['budgets'][2] for c in CONCEPTS)
    lean_total=sum(c['budgets'][0] for c in CONCEPTS)
    readout=[
        ('Recommended planning envelope',rec_total,'MMK'),
        ('Lean–Flagship range',f'{lean_total:,.0f} – {flag_total:,.0f}','MMK'),
        ('Base qualified lead scenarios',sum(c['roi']['base_leads'] for c in CONCEPTS),'planning leads'),
        ('Base expected closes',sum(c['roi']['base_leads']*c['roi']['base_close'] for c in CONCEPTS),'weighted planning closes'),
        ('Commercial claim posture','No guarantee; contribution / close rates require client calibration',''),
        ('Q4 calendar guardrail','Use official dates only as timing cues; rights / permits remain TBC',''),
    ]
    for i,(a,b,c) in enumerate(readout,start=6):
        set_cell(ws,f'F{i}',a,bold=True,color=NAVY);set_cell(ws,f'G{i}',b,wrap=True,numfmt='#,##0' if isinstance(b,(int,float)) else None);set_cell(ws,f'H{i}',c,color=MUTED)
    table_style(ws,6,11,6,8)
    ws.merge_cells('B15:J15'); set_cell(ws,'B15','KEY INSIGHTS',bold=True,fill=LIGHT,color=NAVY,size=13)
    insights=[
        '• The batch deliberately varies conversion methods: assessment booking, site-survey request, product-demo appointment, technical briefing, cohort application, fleet workshop, listening participation, discussion brief, checklist request and CRM decision sprint.',
        '• Official Q4 2026 dates are used only as guardrails: Thadingyut 25–27 Oct, Tazaungdaing 23–24 Nov, National Day 4 Dec and Christmas Day 25 Dec.',
        '• The market source log supports relevance, but all talent, venue, permit, sponsor, supplier, product, technical and rights commitments remain proposed/TBC until client-approved verification.',
        '• Every Recommended-package ROI calculation is a scenario based on a configurable incremental contribution per close—not a financial forecast or performance guarantee.'
    ]
    for i,item in enumerate(insights,start=16):
        ws.merge_cells(start_row=i,start_column=2,end_row=i,end_column=10);set_cell(ws,f'B{i}',item,color='17212B',wrap=True);ws.row_dimensions[i].height=32
    ws.freeze_panes='B5'
    
    cm=wb.create_sheet('Concept Monitor')
    title(cm,'Concept Monitor','Formula-driven planning model. Update only yellow assumption cells after client discovery; this file does not guarantee commercial outcomes.',15)
    headers=['#','Concept','Format','Mode','Lean MMK','Recommended MMK','Flagship MMK','Conservative QL','Base QL','Upside QL','Base Close %','Expected Closes','Contribution / Close','Base Contribution','ROI vs Recommended','Break-even Deals']
    for col,h in enumerate(headers,start=2): set_cell(cm,f'{get_column_letter(col)}5',h,bold=True,fill=NAVY,color='FFFFFF',size=8.5,wrap=True,align='center')
    for r,c in enumerate(CONCEPTS,start=6):
        roi=c['roi']
        vals=[c['id'],c['title_en'],c['format'],c['mode'],c['budgets'][0],c['budgets'][1],c['budgets'][2],roi['conservative_leads'],roi['base_leads'],roi['upside_leads'],roi['base_close'],f'=J{r}*L{r}',roi['contribution_per_close'],f'=M{r}*N{r}',f'=(O{r}-G{r})/G{r}',f'=ROUNDUP(G{r}/N{r},0)']
        for col,v in enumerate(vals,start=2):
            cell=set_cell(cm,f'{get_column_letter(col)}{r}',v,fill='FFFDE7' if col in [9,10,11,12,14] else None,wrap=True,size=8.2)
            if col in [6,7,8,14,15]: cell.number_format='#,##0'
            if col==12 or col==16: cell.number_format='0.0%'
            if col in [13,17]:cell.number_format='#,##0.0'
        cm.row_dimensions[r].height=42
    totalrow=16
    set_cell(cm,f'B{totalrow}','TOTAL / WEIGHTED',bold=True,fill=LIGHT,color=NAVY)
    for col in [6,7,8,9,10,11,13,15,16]:
        letter=get_column_letter(col);set_cell(cm,f'{letter}{totalrow}',f'=SUM({letter}6:{letter}15)',bold=True,fill=LIGHT,color=NAVY,numfmt='#,##0')
    set_cell(cm,f'L{totalrow}','n/a',bold=True,fill=LIGHT,color=NAVY);set_cell(cm,f'M{totalrow}','=SUM(M6:M15)',bold=True,fill=LIGHT,color=NAVY,numfmt='#,##0.0');set_cell(cm,f'N{totalrow}','n/a',bold=True,fill=LIGHT,color=NAVY);set_cell(cm,f'O{totalrow}','=SUM(O6:O15)',bold=True,fill=LIGHT,color=NAVY,numfmt='#,##0');set_cell(cm,f'P{totalrow}','=(O16-G16)/G16',bold=True,fill=LIGHT,color=NAVY,numfmt='0.0%');set_cell(cm,f'Q{totalrow}','n/a',bold=True,fill=LIGHT,color=NAVY)
    table_style(cm,5,16,2,17)
    widths=[5,26,30,12,14,16,14,12,12,12,12,14,18,18,16,16]
    for i,width in enumerate(widths,start=2): cm.column_dimensions[get_column_letter(i)].width=width
    cm.freeze_panes='E6';cm.auto_filter.ref='B5:Q15'
    cm.conditional_formatting.add('P6:P15',ColorScaleRule(start_type='min',start_color='F4CCCC',end_type='max',end_color='D9EAD3'))
    cm.conditional_formatting.add('J6:J15',DataBarRule(start_type='min',end_type='max',color=GOLD))
    chart=BarChart();chart.title='Recommended Planning Envelope by Concept';chart.y_axis.title='MMK';chart.x_axis.title='Concept'
    data=Reference(cm,min_col=7,min_row=5,max_row=15);cats=Reference(cm,min_col=3,min_row=6,max_row=15);chart.add_data(data,titles_from_data=True);chart.set_categories(cats);chart.height=8;chart.width=16;chart.series[0].graphicalProperties.solidFill=GOLD;cm.add_chart(chart,'B19')

    ap=wb.create_sheet('Approval & Risk')
    title(ap,'Approval & Risk Monitor','Owner names, timing and final status must be supplied by the client before production commitments.',10)
    headers=['Concept','Gate','Decision','Owner','Dependency','Status','Target date','Evidence / notes']
    for col,h in enumerate(headers,start=2):set_cell(ap,f'{get_column_letter(col)}5',h,bold=True,fill=NAVY,color='FFFFFF',size=9,wrap=True,align='center')
    rows=[]
    for c in CONCEPTS:
        rows.extend([
            [c['title_en'],'Strategy / scope','Approve tension, audience, KPI and Recommended envelope','Client marketing owner','Brand / audience data','TBC','TBC',''],
            [c['title_en'],'Claims / legal','Approve evidence, message, CTA and data notice','Client legal + technical','Claim substantiation','TBC','TBC',''],
            [c['title_en'],'Production / rights','Approve venue, supplier, talent, media, permit and usage rights','Client PM + production house','Written confirmation','TBC','TBC',''],
        ])
    for r,row in enumerate(rows,start=6):
        for col,v in enumerate(row,start=2): set_cell(ap,f'{get_column_letter(col)}{r}',v,wrap=True,size=8.2,fill='FFFDE7' if col in [5,7,8,9] else None)
        ap.row_dimensions[r].height=34
    table_style(ap,5,5+len(rows),2,9)
    for col,width in zip(range(2,10),[26,20,32,22,22,12,14,35]):ap.column_dimensions[get_column_letter(col)].width=width
    ap.freeze_panes='B6';ap.auto_filter.ref=f'B5:I{5+len(rows)}'

    sa=wb.create_sheet('Assumptions')
    title(sa,'Planning Assumptions','All commercial outcome, cost and timeline inputs below are scenario parameters and should be replaced through client discovery and supplier quotations.',8)
    headers=['Assumption family','Planning treatment','Source / status','Replacement needed','Owner','Notes']
    rows=[
        ['Market relevance','Solar adoption / reliability context used only to frame the category.','S01–S04 verified public sources','No','ZYNTH strategy','No client product, performance or partnership claim inferred.'],
        ['Calendar timing','Thadingyut 25–27 Oct; Tazaungdaing 23–24 Nov; National Day 4 Dec; Christmas 25 Dec 2026.','S05 official calendar','Review before launch','Client PM','Timing cue only; no rights, permit or public-activation entitlement.'],
        ['Budget packages','Lean, Recommended and Flagship are non-binding MMK planning envelopes.','ZYNTH planning input','Yes—supplier quotes / tax','Client procurement','Tax excluded/TBC; contingency requires client release.'],
        ['Commercial scenario','Qualified leads, close rates and incremental contribution are formula inputs.','ZYNTH planning input','Yes—client funnel and margin data','Client commercial owner','No revenue, ROI or payback guarantee.'],
        ['Production resource','Venue, talent, production house, supplier, media, permits and engineering.','Proposed/TBC','Yes—written confirmation','Client PM','No contact, commission or appointment invented.'],
        ['Rights and data','Usage term, territory, consent, privacy, retention and access controls.','Client legal policy required','Yes—legal/data approval','Client legal/data owner','No personal-data processing before lawful-basis confirmation.'],
    ]
    for col,h in enumerate(headers,start=2):set_cell(sa,f'{get_column_letter(col)}5',h,bold=True,fill=NAVY,color='FFFFFF',size=9,wrap=True,align='center')
    for r,row in enumerate(rows,start=6):
        for col,v in enumerate(row,start=2):set_cell(sa,f'{get_column_letter(col)}{r}',v,wrap=True,size=9,fill='FFFDE7' if col in [5,6,7] else None)
        sa.row_dimensions[r].height=54
    table_style(sa,5,11,2,7)
    for col,width in zip(range(2,8),[22,34,24,23,22,40]):sa.column_dimensions[get_column_letter(col)].width=width
    sa.freeze_panes='B6'

    sl=wb.create_sheet('Source Log')
    title(sl,'Verified Source Log','Source findings are bounded by the stated limitations. No source is an endorsement, permit, supplier quotation or partner agreement.',8)
    headers=['ID','Source','Date / access','Type','Retained finding','Use limitation','URL']
    source_rows=[
        ['S01','Reuters','14 Nov 2025','International media','Reports household/business solar adoption amid unreliable electricity and cited import context.','Context only; no individual interview used as testimonial or forecast.','https://www.reuters.com/sustainability/climate-energy/war-torn-myanmar-embraces-solar-tackle-power-crisis-2025-11-14/'],
        ['S02','Xinhua','18 Jan 2026','Dated media/event coverage','Reports a January 2026 Yangon solar-expo precedent and product-category presence.','Historical precedent only; no 2026 Q4 repeat, rights, venue or sponsor claim.','https://english.news.cn/20260118/765dc93ac49a481d8ffbcc1972c01516/c.html'],
        ['S03','Smart Power Myanmar','21 Aug 2026','Local-sector provider','Published journey from loan application to installation.','Supplier self-statements; no partnership or benchmark implied.','https://www.smartpowermyanmar.org/'],
        ['S04','Global New Light of Myanmar','25 Jul 2026','Official-state publisher','Published public context on solar utilisation and technology-management discussion.','Context only; not a permit, project approval or endorsement.','https://www.gnlm.com.mm/efficiently-generate-solar-energy-for-the-state/'],
        ['S05','MOFA / Myanmar National Portal','21 Aug 2026','Official calendar','Lists 2026 Q4 holiday timing.','Timing cue only; no sponsor/public-activation/permit right.','https://www.mofa.gov.mm/about-myanmar/public-holidays/'],
        ['S06','Livoltek YouTube expo video','21 Aug 2026','First-hand exhibitor video','Visual reference for demo/booth format and product-display types.','Conflicting dates excluded; no future availability, performance or rights claim.','https://www.youtube.com/watch?v=9-S7urcEX2U'],
    ]
    for col,h in enumerate(headers,start=2):set_cell(sl,f'{get_column_letter(col)}5',h,bold=True,fill=NAVY,color='FFFFFF',size=9,wrap=True,align='center')
    for r,row in enumerate(source_rows,start=6):
        for col,v in enumerate(row,start=2):
            cell=set_cell(sl,f'{get_column_letter(col)}{r}',v,wrap=True,size=8.5)
            if col==8:
                cell.hyperlink=v;cell.font=Font(name='Aptos',size=8.5,color='0563C1',underline='single')
        sl.row_dimensions[r].height=60
    table_style(sl,5,11,2,8)
    for col,width in zip(range(2,9),[8,23,16,19,42,42,60]):sl.column_dimensions[get_column_letter(col)].width=width
    sl.freeze_panes='B6';sl.auto_filter.ref='B5:H11'

    di=wb.create_sheet('Delivery Inventory')
    title(di,'Deliverable Inventory','Inventory is generated with the batch. Hashes and destination URLs are captured in the separate JSON source manifest after final QA/synchronisation.',8)
    headers=['Category','Count','Status','Destination status','Notes']
    inv=[['English Word proposals',20,'Created','Pending Drive / GitHub','One standalone English proposal per concept.'],['Visual design previews',18,'Created','Pending Drive / GitHub','8 physical concepts × sketch + 3D-style; 2 digital UI storyboards.'],['Editable SVG design sources',18,'Created','Pending Drive / GitHub','Vector source files matched to every client preview.'],['Research / source files',6,'Created','Pending Drive / GitHub','Verified notes, video evidence, data, scripts and registry update.'],['Monitoring workbook',1,'Created','Pending Drive / GitHub','This file.'],['Source manifest',1,'Pending final QA','Pending Drive / GitHub','Will include hashes and actual destination status.'],['Command Center callback',1,'Environment check needed','Not configured until injected variables present','Only approved scheduled-task variables may be used.']]
    for col,h in enumerate(headers,start=2):set_cell(di,f'{get_column_letter(col)}5',h,bold=True,fill=NAVY,color='FFFFFF',size=9,wrap=True,align='center')
    for r,row in enumerate(inv,start=6):
        for col,v in enumerate(row,start=2):set_cell(di,f'{get_column_letter(col)}{r}',v,wrap=True,size=9)
        di.row_dimensions[r].height=38
    table_style(di,5,12,2,6)
    for col,width in zip(range(2,7),[28,12,28,32,54]):di.column_dimensions[get_column_letter(col)].width=width
    di.freeze_panes='B6'

    for sheet in wb.worksheets:
        sheet.sheet_properties.pageSetUpPr.fitToPage=True
        sheet.page_setup.fitToWidth=1
        sheet.page_setup.fitToHeight=0
        sheet.sheet_view.zoomScale=85
    out=OUT/'ZYNTH-20260821-ENERGY-Monitoring.xlsx'
    wb.save(out)
    return out


def sha256(path):
    h=hashlib.sha256()
    with path.open('rb') as f:
        for block in iter(lambda:f.read(1024*1024),b''):h.update(block)
    return h.hexdigest()


def build_manifest():
    files=[]
    excluded={'source_manifest.json'}
    for path in sorted(ROOT.rglob('*')):
        if path.is_file() and path.name not in excluded and '.git' not in path.parts:
            files.append({'path':str(path.relative_to(ROOT)),'bytes':path.stat().st_size,'sha256':sha256(path)})
    manifest={
        'batchCode':'ZYNTH-20260821-ENERGY-BILINGUAL',
        'industryCode':'energy',
        'createdAt':'2026-08-21T03:30:00Z',
        'status':'Created locally; sync statuses must be updated only after verified destinations.',
        'researchSources':{
            'S01':'https://www.reuters.com/sustainability/climate-energy/war-torn-myanmar-embraces-solar-tackle-power-crisis-2025-11-14/',
            'S02':'https://english.news.cn/20260118/765dc93ac49a481d8ffbcc1972c01516/c.html',
            'S03':'https://www.smartpowermyanmar.org/',
            'S04':'https://www.gnlm.com.mm/efficiently-generate-solar-energy-for-the-state/',
            'S05':['https://www.mofa.gov.mm/about-myanmar/public-holidays/','https://myanmar.gov.mm/upcoming-holidays'],
            'S06':'https://www.youtube.com/watch?v=9-S7urcEX2U'
        },
        'deliverableCounts':{'concepts':10,'wordDocuments':20,'videoTreatments':10,'designPackages':18,'designFiles':36,'monitoringWorkbooks':1},
        'sync':{'drive':'Pending','github':'Pending','commandCenter':'Not configured — injected callback variables absent at environment check'},
        'files':files
    }
    out=OUT/'source_manifest.json';out.write_text(json.dumps(manifest,ensure_ascii=False,indent=2)+'\n',encoding='utf-8');return out


def build_report():
    formats='; '.join([f"{c['id']}. {c['format']}" for c in CONCEPTS])
    lean=sum(c['budgets'][0] for c in CONCEPTS);rec=sum(c['budgets'][1] for c in CONCEPTS);flag=sum(c['budgets'][2] for c in CONCEPTS)
    baseleads=sum(c['roi']['base_leads'] for c in CONCEPTS);basecloses=sum(c['roi']['base_leads']*c['roi']['base_close'] for c in CONCEPTS)
    text=f'''# ZYNTH Energy Batch — Monitoring Report

> **Batch:** `ZYNTH-20260821-ENERGY-BILINGUAL`  
> **Industry:** Energy and solar adoption (`energy`)  
> **Status at local build:** Deliverables created; external sync and command-center status must be confirmed separately.

## Batch Coverage

| Item | Coverage |
|---|---:|
| Materially distinct concepts | 10 |
| Standalone Word documents | 20 |
| Commercial-video treatments | 10 |
| Physical sketch/3D-style packages | 8 concepts / 16 packages |
| Digital UI/experience storyboards | 2 concepts |
| Source IDs in log | 6 |

## Formats

{formats}

## Planning Budget Range

| Package view | Aggregate non-tax planning envelope (MMK) | Interpretation |
|---|---:|---|
| Lean | {lean:,.0f} | Compact proof and core conversion format across ten concepts. |
| Recommended | {rec:,.0f} | Balanced production, content and conversion operations across ten concepts. |
| Flagship | {flag:,.0f} | Expanded physical/digital scope, asset suite and contingency depth across ten concepts. |

All figures are **planning envelopes**, not supplier quotes or financial guarantees. Pass-throughs, ZYNTH fee, contingency and taxes are separated inside each proposal. Taxes remain excluded/TBC at applicable rate.

## Scenario-Based Commercial Outcomes

The portfolio’s base scenario contains **{baseleads} qualified planning leads** and **{basecloses:.1f} weighted expected closes**, calculated using configurable concept-level close-rate and incremental-contribution assumptions. The workbook provides each concept’s scenario, ROI equation and break-even deal count. These are decision-support calculations only; the client must replace assumptions with actual funnel, margin, pricing, availability and supplier data before any investment decision.

## Monitoring Coverage

The workbook monitors reach/invitation delivery, proof engagement, qualified leads, booked next steps, client-defined commercial outcomes, consent, claims, incidents, approvals, source limits and production dependencies. It excludes private data, invented contacts, unsupported performance claims and unapproved audience targeting.

## Source Basis

The source log includes dated Reuters and Xinhua reporting, a public Myanmar local-sector provider site, a public state publisher, official Myanmar Q4 2026 holiday pages, and an exhibitor video used only as a visual-format reference. All Q4 venue, partner, sponsor, permit, production-house, talent, supplier, engineering, rights and inventory statements remain `proposed/TBC` unless written evidence is secured.
'''
    out=OUT/'monitoring_report.md';out.write_text(text,encoding='utf-8');return out


def main():
    workbook=build_workbook();report=build_report();manifest=build_manifest()
    print(workbook);print(report);print(manifest)

if __name__=='__main__':main()

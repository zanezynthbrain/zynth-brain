from pathlib import Path
import json
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT = Path('/home/ubuntu/zynth-brain/backend/outputs/zynth_batches/2026-08-21-energy-bilingual/zero_cost_system')
source = ROOT / 'ZYNTH_Zero_Cost_Master_Tracker.xlsx'
out = ROOT / 'native_sheet_values_payload.json'
wb = load_workbook(source, data_only=False)

def cell_value(v):
    if v is None:
        return ''
    if isinstance(v, str) and v.startswith('='):
        return v
    if isinstance(v, (str, int, float, bool)):
        return v
    return str(v)

data = []
for ws in wb.worksheets:
    values = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        values.append([cell_value(cell.value) for cell in row])
    quoted_title = "'" + ws.title.replace("'", "''") + "'"
    rng = f"{quoted_title}!A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    data.append({'range': rng, 'majorDimension': 'ROWS', 'values': values})

payload = {'valueInputOption': 'USER_ENTERED', 'includeValuesInResponse': False, 'data': data}
out.write_text(json.dumps(payload, ensure_ascii=False, separators=(',', ':')), encoding='utf-8')
print(out)
print('Ranges:', len(data), 'Cells:', sum(len(row) for item in data for row in item['values']))

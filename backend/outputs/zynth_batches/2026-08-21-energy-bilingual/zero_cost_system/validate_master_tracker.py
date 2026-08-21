from pathlib import Path
from openpyxl import load_workbook

ROOT = Path('/home/ubuntu/zynth-brain/backend/outputs/zynth_batches/2026-08-21-energy-bilingual/zero_cost_system')
path = ROOT / 'ZYNTH_Zero_Cost_Master_Tracker.xlsx'
wb = load_workbook(path, data_only=False)
errors = []
checks = []

expected_sheets = ['Overview', 'Campaigns', 'Commercials', 'Research & Sources', 'AI Council', 'Ops', 'Learning & Guide']
checks.append(('Workbook sheets', wb.sheetnames == expected_sheets, f'Found: {wb.sheetnames}'))

camp = wb['Campaigns']
com = wb['Commercials']
rs = wb['Research & Sources']
aic = wb['AI Council']
ops = wb['Ops']
guide = wb['Learning & Guide']
ov = wb['Overview']

campaign_ids = [camp.cell(r,2).value for r in range(7,17)]
commercial_ids = [com.cell(r,2).value for r in range(7,17)]
checks.append(('Campaign count = 10', len([x for x in campaign_ids if x]) == 10, str(campaign_ids)))
checks.append(('Commercial count = 10', len([x for x in commercial_ids if x]) == 10, str(commercial_ids)))
checks.append(('Campaign IDs unique', len(set(campaign_ids)) == 10, 'Duplicate IDs checked'))
checks.append(('Commercial IDs unique', len(set(commercial_ids)) == 10, 'Duplicate IDs checked'))
checks.append(('Campaign/commercial IDs distinct', set(campaign_ids).isdisjoint(set(commercial_ids)), 'Separate tracks confirmed'))
checks.append(('Research sources = 6', len([rs.cell(r,2).value for r in range(7,13) if rs.cell(r,2).value]) == 6, 'Source rows 7:12'))
checks.append(('AI council seed records = 10', len([aic.cell(r,2).value for r in range(7,17) if aic.cell(r,2).value]) == 10, 'Contribution IDs 7:16'))
checks.append(('Ops seed records = 6', len([ops.cell(r,2).value for r in range(7,13) if ops.cell(r,2).value]) == 6, 'Ops rows 7:12'))

# Check formula fields on campaign records
formula_ok = True
for r in range(7,17):
    formula_ok &= isinstance(camp.cell(r,18).value, str) and camp.cell(r,18).value.startswith('=')
    formula_ok &= isinstance(camp.cell(r,19).value, str) and camp.cell(r,19).value.startswith('=')
checks.append(('Campaign contribution / ROI formulas present', formula_ok, 'Columns R:S rows 7:16'))

# Validate dropdown references in every data-validation rule
all_dv = []
for ws in [camp, com, rs, aic, ops]:
    all_dv.extend([(ws.title, dv.formula1, str(dv.sqref)) for dv in ws.data_validations.dataValidation])
required_refs = ["='Learning & Guide'!$R$2:$R$9", "='Learning & Guide'!$S$2:$S$9", "='Learning & Guide'!$T$2:$T$5", "='Learning & Guide'!$U$2:$U$6", "='Learning & Guide'!$V$2:$V$5"]
all_formulas = [formula for _, formula, _ in all_dv]
checks.append(('Controlled dropdown rules created', len(all_dv) >= 10, f'{len(all_dv)} rules'))
checks.append(('All dropdown list references available', all(ref in all_formulas for ref in required_refs), str(all_formulas)))

# Ensure links/limitations present
source_links = all(rs.cell(r,7).hyperlink is not None for r in range(7,13))
source_limits = all(bool(rs.cell(r,6).value) for r in range(7,13))
checks.append(('All sources linked', source_links, 'Source URL hyperlinks'))
checks.append(('All sources have limitations', source_limits, 'Use limitation column'))
asset_links = all(camp.cell(r,22).hyperlink is not None for r in range(7,17)) and all(com.cell(r,24).hyperlink is not None for r in range(7,17))
checks.append(('All current campaign/commercial records linked to assets', asset_links, 'GitHub batch links'))

# Guide and dashboard
checks.append(('Zero-cost guide exists', 'API မလိုပါ' in str(guide['B3'].value), str(guide['B3'].value)))
checks.append(('Overview formula tiles exist', all(isinstance(ov.cell(r,c).value, str) and ov.cell(r,c).value.startswith('=') for r,c in [(8,2),(8,5),(8,8),(12,2),(12,5),(12,8)]), 'Overview summary formulas'))

for name, ok, detail in checks:
    if not ok:
        errors.append(f'- FAIL: {name} — {detail}')

status = 'PASS' if not errors else 'FAIL'
report = [
    '# ZYNTH Zero-Cost Master Tracker — Validation',
    '',
    f'**Status:** {status}',
    '',
    '| Check | Result | Detail |',
    '|---|---|---|',
]
for name, ok, detail in checks:
    report.append(f'| {name} | {"PASS" if ok else "FAIL"} | {detail.replace("|", "/")} |')
report.extend([
    '',
    '## Validation Notes',
    '',
    '- The workbook contains formulas that calculate when opened in Microsoft Excel or Google Sheets; calculation mode is set to automatic.',
    '- The future Looker Studio dashboard should connect to the tabular record tabs after this one workbook is uploaded to Google Drive and opened as Google Sheets.',
    '- The initial solution contains no paid API key, provider token, paid automation or custom server.',
])
(ROOT / 'master_tracker_validation.md').write_text('\n'.join(report) + '\n', encoding='utf-8')
print(status)
print('\n'.join(errors) if errors else 'All checks passed.')

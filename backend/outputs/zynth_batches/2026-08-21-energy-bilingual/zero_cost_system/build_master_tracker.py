from pathlib import Path
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.chart import DoughnutChart, BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

ROOT = Path('/home/ubuntu/zynth-brain/backend/outputs/zynth_batches/2026-08-21-energy-bilingual')
OUT = ROOT / 'zero_cost_system' / 'ZYNTH_Zero_Cost_Master_Tracker.xlsx'
CONCEPTS = json.loads((ROOT / 'data' / 'concepts.json').read_text(encoding='utf-8'))

# Theme
NAVY = '17324D'
TEAL = '1A5F5F'
LIGHT_BLUE = 'E6F3FF'
LIGHT_TEAL = 'E8F5F2'
LIGHT_YELLOW = 'FFFDE7'
LIGHT_ORANGE = 'FFF3E0'
LIGHT_GREEN = 'E8F5E9'
LIGHT_RED = 'FFCCBC'
LIGHT_GRAY = 'F2F4F7'
MID_GRAY = '667085'
WHITE = 'FFFFFF'
BORDER = 'D0D5DD'
PURPLE = '4A235A'
FONT = 'Noto Sans Myanmar'
FONT_FALLBACK = 'Calibri'

thin = Side(style='thin', color=BORDER)
medium = Side(style='medium', color=NAVY)

wb = Workbook()
wb.remove(wb.active)
for name in ['Overview', 'Campaigns', 'Commercials', 'Research & Sources', 'AI Council', 'Ops', 'Learning & Guide']:
    wb.create_sheet(name)
wb.calculation.fullCalcOnLoad = True
wb.calculation.forceFullCalc = True
wb.calculation.calcMode = 'auto'


def style_title(ws, title, subtitle, last_col):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=last_col)
    c = ws.cell(2, 2, title)
    c.font = Font(name=FONT, size=20, bold=True, color=NAVY)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 30
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=last_col)
    c = ws.cell(3, 2, subtitle)
    c.font = Font(name=FONT, size=10, italic=True, color=MID_GRAY)
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[3].height = 26


def section(ws, row, title, last_col, color=LIGHT_BLUE):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=last_col)
    c = ws.cell(row, 2, title)
    c.font = Font(name=FONT, size=12, bold=True, color=NAVY)
    c.fill = PatternFill('solid', fgColor=color)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row].height = 23


def header_row(ws, row, headers, start_col=2):
    for i, h in enumerate(headers, start_col):
        c = ws.cell(row, i, h)
        c.font = Font(name=FONT, size=10, bold=True, color=WHITE)
        c.fill = PatternFill('solid', fgColor=NAVY)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = Border(left=thin, right=thin, top=thin, bottom=medium)
    ws.row_dimensions[row].height = 36


def style_table(ws, start_row, end_row, start_col, end_col, wrap_cols=None, shade_every_other=True):
    wrap_cols = set(wrap_cols or [])
    for r in range(start_row, end_row + 1):
        if shade_every_other and (r - start_row) % 2 == 1:
            fill = PatternFill('solid', fgColor='FAFBFC')
        else:
            fill = PatternFill('solid', fgColor=WHITE)
        for col in range(start_col, end_col + 1):
            c = ws.cell(r, col)
            c.font = Font(name=FONT, size=10, color='101828')
            c.fill = fill
            c.border = Border(left=thin if col == start_col else Side(style=None),
                              right=thin if col == end_col else Side(style=None),
                              top=thin, bottom=thin)
            c.alignment = Alignment(horizontal='left' if col in wrap_cols else 'center',
                                    vertical='center', wrap_text=(col in wrap_cols), indent=1 if col in wrap_cols else 0)
        ws.row_dimensions[r].height = 38 if wrap_cols else 22


def widths(ws, mapping):
    for col, width in mapping.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def add_hyperlink(cell, url, label=None):
    cell.value = label or url
    cell.hyperlink = url
    cell.font = Font(name=FONT, size=10, color='0563C1', underline='single')
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)


def add_filter_and_freeze(ws, header_row_num, end_row, end_col):
    ws.auto_filter.ref = f'B{header_row_num}:{get_column_letter(end_col)}{end_row}'
    ws.freeze_panes = f'B{header_row_num + 1}'

# Lists and Guide -----------------------------------------------------------
guide = wb['Learning & Guide']
style_title(guide, 'ZYNTH Master Tracker — အသုံးပြုနည်းနှင့် Learning Library',
            'ဒီ file တစ်ခုတည်းကို အမြဲ update လုပ်ပါ။ Proposal အသစ်တိုင်းအတွက် workbook အသစ်မဖန်တီးပါနှင့်။ API မလိုပါ။', 12)
section(guide, 5, 'START HERE — 7-STEP UPDATE ROUTINE', 12, LIGHT_TEAL)
steps = [
    ('1', 'Campaign / Commercial ID အသစ်သတ်မှတ်ပါ', 'ဥပမာ CMP-2026-ENERGY-011 သို့မဟုတ် COM-2026-ENERGY-011။ ID ကို မပြောင်းပါနှင့်။'),
    ('2', 'Research & Sources တွင် source/claim ကိုအရင်မှတ်တမ်းတင်ပါ', 'Claim, URL, access date, limitation နှင့် confidence ကိုထည့်ပါ။'),
    ('3', 'Campaigns နှင့် Commercials တွင် row အသစ်ထည့်ပါ', 'Campaign 10 ခုနှင့် standalone commercial/storyboard 10 ခုကို သီးခြားနံပါတ်ပေးပါ။'),
    ('4', 'Manual AI Council packet ကို paste လုပ်ပါ', 'AI web interface တိုင်းသို့ canonical brief တစ်ခုတည်းပို့ပြီး structured output ကို AI Council tab တွင်ထည့်ပါ။'),
    ('5', 'G0–G7 gate နှင့် approval ကို Ops tab တွင်ပြောင်းပါ', 'Status dropdown ကိုသာအသုံးပြုပါ။ Free-text status မရေးပါနှင့်။'),
    ('6', 'Artifact link ကို update လုပ်ပါ', 'Drive/GitHub မှ latest proposal, treatment, storyboard ကို link ချိတ်ပါ။ Current Version ကို material change တစ်ခုစီတိုင်းတိုးပါ။'),
    ('7', 'Overview နှင့် Looker Studio ကိုစစ်ပါ', 'Sheet ကို Google Sheets သို့တင်ပြီး Looker Studio ကို Overview-ready tab များနှင့်ချိတ်ပါ။')
]
header_row(guide, 6, ['Step', 'ဘာလုပ်မလဲ', 'ဘာကြောင့်အရေးကြီးလဲ'], 2)
for idx, (n, a, b) in enumerate(steps, 7):
    guide.cell(idx, 2, n)
    guide.cell(idx, 3, a)
    guide.cell(idx, 4, b)
style_table(guide, 7, 13, 2, 4, wrap_cols=[3,4])
widths(guide, {2:10,3:36,4:78,5:4,6:20,7:20,8:20,9:20,10:20,11:20,12:20})
section(guide, 16, 'THE 10 QUESTIONS TO ASK OF ANY PROPOSAL', 12, LIGHT_YELLOW)
questions = [
    'What human/commercial tension is actually being solved?',
    'Which audience is most likely to change behaviour, and what evidence supports that?',
    'What is the one behaviour or decision we want next?',
    'Is the conversion mechanism realistic and consented?',
    'Which assertions are facts, which are assumptions, and which are proposed/TBC?',
    'What makes this creative route distinct from the other nine?',
    'Can it be produced safely, legally, and on the proposed timing/budget?',
    'What must be true for the ROI scenario to work?',
    'Who owns the next step, and what happens if they do not act?',
    'What will we learn even if the result is weaker than expected?'
]
header_row(guide, 17, ['#', 'Review question'], 2)
for r, q in enumerate(questions, 18):
    guide.cell(r, 2, r-17)
    guide.cell(r, 3, q)
style_table(guide, 18, 27, 2, 3, wrap_cols=[3])
section(guide, 30, 'ZERO-COST-FIRST RULES', 12, LIGHT_ORANGE)
rules = [
    'Do not use a paid API, paid automation or new subscription until the team approves a measurable upgrade case.',
    'Use existing AI web interfaces manually with the same structured brief; paste the result into AI Council. Do not paste sensitive client data.',
    'Use Google Sheets as the current master record, Drive/GitHub for assets, Looker Studio for dashboarding and NotebookLM/Obsidian only as learning aids.',
    'No AI or automation may approve claims, publish work, contact people, spend money or commit vendors.'
]
for r, rule in enumerate(rules, 31):
    guide.merge_cells(start_row=r, start_column=2, end_row=r, end_column=12)
    c = guide.cell(r,2, '• ' + rule)
    c.font = Font(name=FONT, size=10, color='344054')
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    c.fill = PatternFill('solid', fgColor='FFF9F2')
    ws_border = Border(left=thin,right=thin,top=thin,bottom=thin)
    c.border = ws_border
    guide.row_dimensions[r].height = 28

# Lists in hidden area
list_cols = {
    18: ('Status', ['Draft','In Review','Needs Human Decision','Approved','On Hold','Live','Closed','Failed']),
    19: ('Stage', ['G0 Brief Accepted','G1 Research Cleared','G2 Strategy Selected','G3 Creative Selected','G4 Feasibility Cleared','G5 Go / No-go','G6 Live Optimisation','G7 Learn & Archive']),
    20: ('Confidence', ['High','Medium','Low','Needs review']),
    21: ('Decision', ['Pending','Accept','Revise','Hold','Reject']),
    22: ('Priority', ['Critical','High','Medium','Low']),
}
for col, (label, values) in list_cols.items():
    guide.cell(1, col, label)
    for r, value in enumerate(values,2):
        guide.cell(r,col,value)
    guide.column_dimensions[get_column_letter(col)].hidden = True

guide.sheet_view.showGridLines = False
guide.freeze_panes = 'B7'

# Overview -----------------------------------------------------------------
ov = wb['Overview']
style_title(ov, 'ZYNTH Command Center — Zero-Cost-First',
            'One living proposal portfolio. Update the records below; do not make a new workbook for every batch. Generated 2026-08-21 UTC.', 12)
section(ov, 5, 'PORTFOLIO SNAPSHOT', 12, LIGHT_TEAL)
metrics = [
    ('Campaign concepts', '=COUNTA(Campaigns!B7:B200)', 'Campaigns'),
    ('Commercial concepts', '=COUNTA(Commercials!B7:B200)', 'Standalone commercial/storyboard seeds'),
    ('Research claims', '=COUNTA(\'Research & Sources\'!B7:B200)', 'Verified/context claims logged'),
    ('Pending AI decisions', '=COUNTIF(\'AI Council\'!K7:K200,"Pending")', 'Human decision required'),
    ('Open / at-risk tasks', '=COUNTIFS(Ops!J7:J200,"<>Closed",Ops!J7:J200,"<>Completed")', 'Operational follow-through'),
    ('Approved / live campaigns', '=COUNTIF(Campaigns!J7:J200,"Approved")+COUNTIF(Campaigns!J7:J200,"Live")', 'Ready or currently active'),
]
for i, (label, formula, note) in enumerate(metrics):
    row = 7 + (i//3)*4
    col = 2 + (i%3)*3
    ov.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
    c = ov.cell(row,col,label)
    c.font = Font(name=FONT, size=10, bold=True, color=MID_GRAY)
    c.fill = PatternFill('solid', fgColor=LIGHT_BLUE)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ov.merge_cells(start_row=row+1,start_column=col,end_row=row+2,end_column=col+1)
    c = ov.cell(row+1,col,formula)
    c.font = Font(name=FONT, size=22, bold=True, color=NAVY)
    c.fill = PatternFill('solid', fgColor=WHITE)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = Border(left=thin,right=thin,top=thin,bottom=thin)
    ov.merge_cells(start_row=row+3,start_column=col,end_row=row+3,end_column=col+1)
    c = ov.cell(row+3,col,note)
    c.font = Font(name=FONT, size=9, italic=True, color=MID_GRAY)
    c.fill = PatternFill('solid', fgColor=WHITE)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for rr in range(row,row+4):
        for cc in range(col,col+2):
            ov.cell(rr,cc).border = Border(left=thin,right=thin,top=thin,bottom=thin)
    ov.row_dimensions[row+1].height = 24
    ov.row_dimensions[row+2].height = 24
section(ov, 17, 'CURRENT BATCH — WHAT REQUIRES A DECISION?', 12, LIGHT_YELLOW)
header_row(ov, 18, ['Track','Record ID','Title','Current Stage','Status','Next Gate','Owner','Recommended Budget (MMK)','Base ROI','Latest Asset'],2)
for i, concept in enumerate(CONCEPTS,19):
    values = [
        'Campaign', f"CMP-2026-ENERGY-{concept['id']}", concept['title_en'], 'G3 Creative Selected', 'In Review',
        'Client route selection / detailed feasibility', 'ZYNTH Strategy Lead', concept['budgets'][1],
        f"=IFERROR((Campaigns!R{7+i-19}-Campaigns!N{7+i-19})/Campaigns!N{7+i-19},\"\")", 'Open Campaigns tab'
    ]
    for j,v in enumerate(values,2): ov.cell(i,j,v)
    ov.cell(i,9).number_format = '#,##0'
    ov.cell(i,10).number_format = '0.0%'
    ov.row_dimensions[i].height=24
style_table(ov,19,18+len(CONCEPTS),2,11,wrap_cols=[4,5,7,8,11])
add_filter_and_freeze(ov,18,18+len(CONCEPTS),11)
widths(ov,{2:14,3:23,4:34,5:21,6:18,7:34,8:21,9:18,10:12,11:24,12:4})
for row in range(19,19+len(CONCEPTS)):
    ov.cell(row,9).number_format = '#,##0'
    ov.cell(row,10).number_format = '0.0%'

# Campaigns ----------------------------------------------------------------
camp = wb['Campaigns']
style_title(camp, 'Campaign Portfolio — Campaign / Event / Activation Track',
            'One row = one campaign concept. Update this same ID across every proposal cycle. Planning values are assumptions, not quotations or guarantees.', 23)
section(camp,5,'CAMPAIGN MASTER RECORDS',23,LIGHT_TEAL)
camp_headers = ['Record ID','Campaign Title (EN)','Campaign Title (MY)','Format','Commercial Tension','Audience / Behaviour Change','Conversion Mechanic','Creative Territory','Current Stage','Status','Owner','Next Gate','Recommended Budget (MMK)','Base Qualified Leads','Base Close Rate','Contribution / Close (MMK)','Base Contribution (MMK)','Base ROI','Current Version','Research Confidence','Latest Artifact / Folder','Last Updated']
header_row(camp,6,camp_headers,2)
base_url='https://github.com/zanezynthbrain/zynth-brain/tree/main/backend/outputs/zynth_batches/2026-08-21-energy-bilingual'
for r, concept in enumerate(CONCEPTS,7):
    values = [
        f"CMP-2026-ENERGY-{concept['id']}", concept['title_en'], concept['title_my'], concept['format'],
        concept['commercial_tension'], concept['behaviour_change'], concept['conversion'], concept['territory'],
        'G3 Creative Selected', 'In Review', 'ZYNTH Strategy Lead', 'Client route selection / detailed feasibility',
        concept['budgets'][1], concept['roi']['base_leads'], concept['roi']['base_close'], concept['roi']['contribution_per_close'],
        f'=N{r}*O{r}*P{r}', f'=IFERROR((Q{r}-N{r})/N{r},"")', 1, 'Medium', 'Batch folder', '2026-08-21 UTC'
    ]
    for c,v in enumerate(values,2): camp.cell(r,c,v)
    add_hyperlink(camp.cell(r,22),base_url,'Open batch')
style_table(camp,7,6+len(CONCEPTS),2,23,wrap_cols=[3,4,5,6,7,8,9,12,22,23])
widths(camp,{2:22,3:34,4:30,5:32,6:42,7:36,8:33,9:24,10:22,11:18,12:20,13:34,14:19,15:18,16:16,17:21,18:22,19:14,20:15,21:19,22:24,23:18})
for r in range(7,7+len(CONCEPTS)):
    for col in [14,17,18]: camp.cell(r,col).number_format = '#,##0'
    camp.cell(r,16).number_format = '0.0%'
    camp.cell(r,19).number_format = '0.0%'
    camp.row_dimensions[r].height = 54
add_filter_and_freeze(camp,6,6+len(CONCEPTS),23)

# Commercials --------------------------------------------------------------
com = wb['Commercials']
style_title(com, 'Commercial Studio — Standalone Commercial & Storyboard Track',
            'One row = one independently numbered commercial concept. Current rows are seed treatments from the Energy batch and must be expanded into detailed storyboards before production.', 24)
section(com,5,'COMMERCIAL / STORYBOARD MASTER RECORDS',24,LIGHT_ORANGE)
com_headers = ['Record ID','Parent Campaign ID','Commercial Title (EN)','Commercial Title (MY)','Format','Single-Minded Proposition','Logline / Premise','Visual Style','Story Location','Treatment Status','Storyboard Status','Production Status','Rights / Claims','Current Stage','Status','Owner','Next Gate','Master Film','Social Cutdowns','Recommended Production Budget (MMK)','Current Version','Research Confidence','Latest Artifact / Folder','Last Updated']
header_row(com,6,com_headers,2)
for r, concept in enumerate(CONCEPTS,7):
    values=[
        f"COM-2026-ENERGY-{concept['id']}", f"CMP-2026-ENERGY-{concept['id']}", f"{concept['title_en']} — Commercial Seed", concept['title_my'],
        'Brand / product film seed', concept['message'], concept['video_premise'], concept['visual_style'], concept['story_location'],
        'Seed treatment complete', 'Detailed storyboard required', 'Not started', 'All locations, talent, claims, music and rights TBC',
        'G3 Creative Selected', 'Needs Human Decision', 'ZYNTH Creative Lead', 'Build 12-frame detailed storyboard + production feasibility',
        'TBC', 'TBC', 'TBC', 1, 'Medium', 'Open batch'
        ,'2026-08-21 UTC'
    ]
    for c,v in enumerate(values,2): com.cell(r,c,v)
    add_hyperlink(com.cell(r,24),base_url,'Open batch')
style_table(com,7,6+len(CONCEPTS),2,25,wrap_cols=[4,5,6,7,8,9,10,11,12,13,14,17,18,19,20,24,25])
widths(com,{2:22,3:22,4:34,5:30,6:24,7:36,8:42,9:34,10:30,11:19,12:23,13:18,14:38,15:21,16:21,17:21,18:34,19:14,20:16,21:22,22:14,23:16,24:22,25:18})
for r in range(7,7+len(CONCEPTS)): com.row_dimensions[r].height=62
add_filter_and_freeze(com,6,6+len(CONCEPTS),25)

# Research & Sources ------------------------------------------------------
rs = wb['Research & Sources']
style_title(rs,'Research & Sources — Claim-Level Evidence Register',
            'Every client-facing claim must link to a source, access date, limitation and confidence. A source does not equal an endorsement.',16)
section(rs,5,'VERIFIED SOURCE REGISTER',16,LIGHT_YELLOW)
rs_headers=['Source ID','Publisher / Source','Type','Finding Retained','Use Limitation','Source URL','Access / Publication Date','Confidence','Reviewer','Status']
header_row(rs,6,rs_headers,2)
sources=[
('S01','Reuters','International media','Reports solar adoption amid unreliable electricity and cites Chinese customs data for 2025 solar-panel imports.','Context only; do not use interviews as testimonials or forecasts.','https://www.reuters.com/sustainability/climate-energy/war-torn-myanmar-embraces-solar-tackle-power-crisis-2025-11-14/','14 Nov 2025','Medium','ZYNTH Research','Cleared'),
('S02','Xinhua','Dated media / event coverage','Reports the 16–18 Jan 2026 Myanmar Power and Solar Energy Storage Lighting Expo in Yangon with 70+ exhibitors and 77 brands.','Historical precedent only; no sponsorship, repeat edition or Q4 event assumed.','https://english.news.cn/20260118/765dc93ac49a481d8ffbcc1972c01516/c.html','18 Jan 2026','Medium','ZYNTH Research','Cleared'),
('S03','Smart Power Myanmar','Local-sector provider website','Published journey runs from loan application to installation; shows bank/business logos and individual case studies.','Supplier statements only; do not generalise outcomes or imply partnership.','https://www.smartpowermyanmar.org/','21 Aug 2026','Low','ZYNTH Research','Use with limitation'),
('S04','Global New Light of Myanmar','Official-state publisher','States solar systems are being encouraged and mentions a 15 July solar workshop.','Official-policy context only; not a permit, endorsement or procurement right.','https://www.gnlm.com.mm/efficiently-generate-solar-energy-for-the-state/','25 Jul 2026','Medium','ZYNTH Research','Cleared'),
('S05','Myanmar MOFA / National Portal','Official public calendar','Lists 2026 Thadingyut, Tazaungdaing, National Day and Christmas holidays.','Timing cue only; no activation permit, association or sponsorship implied.','https://www.mofa.gov.mm/about-myanmar/public-holidays/','21 Aug 2026','High','ZYNTH Research','Cleared'),
('S06','Livoltek YouTube expo video','First-hand exhibitor video','Shows booth format, demo, meeting tables and installer/reseller-oriented materials.','Visual format only; conflicting dates/technical claims excluded from factual planning.','https://www.youtube.com/watch?v=9-S7urcEX2U','21 Aug 2026','Low','ZYNTH Research','Use with limitation')]
for r,row in enumerate(sources,7):
    for c,v in enumerate(row,2): rs.cell(r,c,v)
    add_hyperlink(rs.cell(r,7),row[5],'Open source')
style_table(rs,7,12,2,11,wrap_cols=[3,4,5,6,7,8])
widths(rs,{2:12,3:28,4:22,5:45,6:50,7:22,8:18,9:16,10:20,11:22,12:16})
for r in range(7,13): rs.row_dimensions[r].height=70
add_filter_and_freeze(rs,6,12,11)

# AI Council ---------------------------------------------------------------
aic = wb['AI Council']
style_title(aic,'AI Council — Manual, Provider-Neutral Collaboration Log',
            'No API required. Use the same canonical brief in any available AI web interface, then paste only the structured artifact and decision here.',17)
section(aic,5,'AI CONTRIBUTION & HUMAN RESOLUTION REGISTER',17,LIGHT_ORANGE)
aic_headers=['Contribution ID','Record ID','AI Role','Provider / Tool','Brief Version','Contribution Summary / Artifact Link','Evidence Count','Pros / Strong Points','Cons / Risks','Critic Result','Human Decision','Decision Owner','Resolution / Next Action','Created At']
header_row(aic,6,aic_headers,2)
roles=['Research','Strategy','Creative','Production','Critic','Research','Strategy','Creative','Production','Critic']
for r, concept in enumerate(CONCEPTS,7):
    values=[f"AIC-2026-ENERGY-{concept['id']}",f"CMP-2026-ENERGY-{concept['id']}",roles[r-7],'Manual web UI — choose approved tool','v1',
            'Paste structured AI artifact or Drive/GitHub link here.',0,'Pending manual contribution','Pending critique','Pending','Pending','ZYNTH Decision Owner','Select: Accept / Revise / Hold / Reject','2026-08-21 UTC']
    for c,v in enumerate(values,2): aic.cell(r,c,v)
style_table(aic,7,6+len(CONCEPTS),2,15,wrap_cols=[4,5,6,7,9,10,11,14,15])
widths(aic,{2:20,3:22,4:16,5:28,6:13,7:45,8:14,9:32,10:32,11:17,12:17,13:23,14:42,15:18})
for r in range(7,7+len(CONCEPTS)): aic.row_dimensions[r].height=50
add_filter_and_freeze(aic,6,6+len(CONCEPTS),15)

# Ops ----------------------------------------------------------------------
ops = wb['Ops']
style_title(ops,'Operations — Tasks, Approvals, Budget & Exceptions',
            'Use this tab for G0–G7 decisions, accountable owners, blockers and approved changes. No external action proceeds without a named human decision.',18)
section(ops,5,'TASKS & APPROVALS',18,LIGHT_TEAL)
ops_headers=['Task / Approval ID','Record ID','Type','Gate','Task / Decision','Accountable Owner','Due Date','Priority','Dependency / Risk','Status','Decision / Condition','Evidence / Asset Link','Last Updated']
header_row(ops,6,ops_headers,2)
ops_rows=[
('TSK-2026-001','CMP-2026-ENERGY-01','Task','G3 Creative Selected','Select campaign route for first pilot','User / Final Approver','TBC','High','Client brief confirmation','Needs Human Decision','TBC','Batch proposal','2026-08-21 UTC'),
('TSK-2026-002','COM-2026-ENERGY-01','Task','G3 Creative Selected','Create 12-frame detailed storyboard','ZYNTH Creative Lead','TBC','High','Creative territory selection','Draft','TBC','Batch treatment','2026-08-21 UTC'),
('APR-2026-001','CMP-2026-ENERGY-01','Approval','G4 Feasibility Cleared','Approve claims / data / rights review before execution','Client approver','TBC','Critical','Evidence, legal, privacy and safety review','Needs Human Decision','TBC','Research register','2026-08-21 UTC'),
('TSK-2026-003','CMP-2026-ENERGY-08','Task','G2 Strategy Selected','Confirm consent language and destination experience','Client data owner','TBC','Critical','Data/privacy policy','Draft','TBC','Proposal 08','2026-08-21 UTC'),
('TSK-2026-004','COM-2026-ENERGY-08','Task','G4 Feasibility Cleared','Define production scope, cutdowns and usage rights','ZYNTH Production Lead','TBC','High','Script and budget approval','Draft','TBC','Commercial seed','2026-08-21 UTC'),
('TSK-2026-005','SYSTEM-ZYNTH','Task','G0 Brief Accepted','Upload this tracker to Drive; assign one data owner','User / Data Owner','TBC','High','Master file owner','Draft','TBC','This workbook','2026-08-21 UTC')]
for r,row in enumerate(ops_rows,7):
    for c,v in enumerate(row,2): ops.cell(r,c,v)
style_table(ops,7,6+len(ops_rows),2,14,wrap_cols=[4,6,7,10,12,13])
widths(ops,{2:20,3:22,4:14,5:22,6:45,7:24,8:15,9:14,10:36,11:22,12:36,13:28,14:18})
for r in range(7,7+len(ops_rows)): ops.row_dimensions[r].height=50
add_filter_and_freeze(ops,6,6+len(ops_rows),14)

# Data validation ----------------------------------------------------------
status_formula = "='Learning & Guide'!$R$2:$R$9"
stage_formula = "='Learning & Guide'!$S$2:$S$9"
confidence_formula = "='Learning & Guide'!$T$2:$T$5"
decision_formula = "='Learning & Guide'!$U$2:$U$6"
priority_formula = "='Learning & Guide'!$V$2:$V$5"
for ws, ranges in [
    (camp, [(10, f'K7:K500', status_formula),(9, f'J7:J500', stage_formula),(21,f'U7:U500', confidence_formula)]),
    (com, [(16, f'P7:P500', status_formula),(15,f'O7:O500',stage_formula),(23,f'W7:W500',confidence_formula)]),
    (rs, [(9,f'I7:I500',confidence_formula),(11,f'K7:K500',status_formula)]),
    (aic, [(11,f'K7:K500',decision_formula),(12,f'L7:L500',decision_formula)]),
    (ops, [(11,f'K7:K500',status_formula),(9,f'I7:I500',priority_formula)])
]:
    for _, cell_range, formula in ranges:
        dv = DataValidation(type='list', formula1=formula, allow_blank=True)
        dv.error = 'Please use the approved list value.'
        dv.errorTitle = 'Controlled field'
        ws.add_data_validation(dv)
        dv.add(cell_range)

# Conditional formats
for ws, status_range in [(camp,'K7:K500'),(com,'P7:P500'),(ops,'K7:K500')]:
    ws.conditional_formatting.add(status_range, FormulaRule(formula=[f'{status_range.split(":")[0]}="Needs Human Decision"'], fill=PatternFill('solid',fgColor=LIGHT_RED)))
    ws.conditional_formatting.add(status_range, FormulaRule(formula=[f'{status_range.split(":")[0]}="Approved"'], fill=PatternFill('solid',fgColor=LIGHT_GREEN)))
    ws.conditional_formatting.add(status_range, FormulaRule(formula=[f'{status_range.split(":")[0]}="Live"'], fill=PatternFill('solid',fgColor=LIGHT_TEAL)))
camp.conditional_formatting.add('S7:S500', ColorScaleRule(start_type='min', start_color='FFCCBC', mid_type='percentile', mid_value=50, mid_color='FFF9C4', end_type='max', end_color='E8F5E9'))

# Comments and input cues
for ws in [camp, com, rs, aic, ops]:
    ws['B6'].comment = Comment('Rows are canonical records. Add new rows under the table and preserve the existing Record ID.', 'ZYNTH')

# Charts on Overview. Openpyxl charts reference formulas only after Excel/Sheets calculates.
chart = DoughnutChart()
chart.title = 'Portfolio by Track'
chart.width = 8
chart.height = 6
ov['B32']='Track'; ov['C32']='Count'
ov['B33']='Campaigns'; ov['C33']='=COUNTA(Campaigns!B7:B200)'
ov['B34']='Commercials'; ov['C34']='=COUNTA(Commercials!B7:B200)'
data=Reference(ov,min_col=3,min_row=32,max_row=34)
labels=Reference(ov,min_col=2,min_row=33,max_row=34)
chart.add_data(data,titles_from_data=True); chart.set_categories(labels)
chart.holeSize=55
ov.add_chart(chart,'B36')
bar = BarChart()
bar.title = 'Recommended Campaign Budget (MMK)'
bar.y_axis.title = 'MMK'
bar.x_axis.title = 'Concept'
bar.width=17
bar.height=7
bar_data=Reference(camp,min_col=14,min_row=6,max_row=16)
bar_cats=Reference(camp,min_col=2,min_row=7,max_row=16)
bar.add_data(bar_data,titles_from_data=True); bar.set_categories(bar_cats)
bar.series[0].graphicalProperties.solidFill=TEAL
ov.add_chart(bar,'H36')

# Tab colors
for ws in [ov,camp,com,rs,aic,ops,guide]:
    ws.sheet_properties.tabColor = TEAL if ws.title in ['Overview','Campaigns','Ops'] else PURPLE

# Footer notes
for ws in [ov,camp,com,rs,aic,ops,guide]:
    ws.oddFooter.center.text = 'ZYNTH Master Tracker | Zero-cost-first | Generated 2026-08-21'
    ws.oddFooter.center.size = 8

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(OUT)
print('Sheets:', ', '.join(wb.sheetnames))
print('Campaigns:', len(CONCEPTS), 'Commercials:', len(CONCEPTS), 'Sources:', len(sources))

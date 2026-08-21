from pathlib import Path
import json, hashlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation

ROOT=Path('/home/ubuntu/zynth-brain/backend/outputs/zynth_batches/2026-08-21-beauty-bilingual'); MON=ROOT/'monitoring'; MON.mkdir(exist_ok=True)
CAM=json.loads((ROOT/'data/campaigns.json').read_text(encoding='utf-8'))['campaigns']; COM=json.loads((ROOT/'data/commercials.json').read_text(encoding='utf-8'))['commercials']
BLACK='2D2D2D'; PLUM='722F37'; ROSE='F4E5E8'; GOLD='B28A4A'; INK='1F2933'; GREY='6B7280'; WHITE='FFFFFF'; thin=Side(style='thin',color='D9D2D5')
def style(ws):
    ws.sheet_view.showGridLines=False; ws.column_dimensions['A'].width=3
    for row in ws.iter_rows():
        for c in row: c.alignment=Alignment(vertical='center',wrap_text=True); c.font=Font(name='Calibri',size=9,color=INK)
def title(ws,t,sub,last):
    ws.merge_cells(start_row=2,start_column=2,end_row=2,end_column=last); ws.cell(2,2,t).font=Font(name='Georgia',size=20,bold=True,color=BLACK); ws.row_dimensions[2].height=30
    ws.merge_cells(start_row=3,start_column=2,end_row=3,end_column=last); ws.cell(3,2,sub).font=Font(name='Calibri',size=9,italic=True,color=GREY); ws.row_dimensions[3].height=28
def grid(ws,row,heads,rows,widths):
    for i,h in enumerate(heads,2):
        c=ws.cell(row,i,h); c.fill=PatternFill('solid',fgColor=BLACK); c.font=Font(name='Calibri',size=9,bold=True,color=WHITE); c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); c.border=Border(left=thin,right=thin,top=thin,bottom=thin)
    ws.row_dimensions[row].height=30
    for rnum,record in enumerate(rows,row+1):
        for i,v in enumerate(record,2):
            c=ws.cell(rnum,i,v); c.border=Border(left=thin,right=thin,top=thin,bottom=thin); c.alignment=Alignment(vertical='center',wrap_text=True)
            if (rnum-row)%2: c.fill=PatternFill('solid',fgColor='FCFAFB')
        ws.row_dimensions[rnum].height=46
    from openpyxl.utils import get_column_letter
    for i,w in enumerate(widths,2): ws.column_dimensions[get_column_letter(i)].width=w
    return row+1,row+len(rows)
def mmk(x): return f'MMK {x/1e6:.1f}m'
wb=Workbook(); ov=wb.active; ov.title='Overview'; style(ov); title(ov,'ZYNTH Beauty / Cosmetics — Command View','Batch: ZYNTH-20260821-BEAUTY-BILINGUAL | Social media and TikTok are campaign preparation workstreams, not substitutes for the separate commercial track.',9)
metrics=[('Campaign proposals','=COUNTA(Campaigns!B7:B16)','Integrated activation + social/TikTok preparation'),('Social/TikTok routes','=COUNTA(Social_TikTok!B7:B16)','One required workstream per campaign'),('Standalone commercials','=COUNTA(Commercials!B7:B16)','Separate COM IDs and decisions'),('Storyboard frames','=SUM(Commercials!L7:L16)','12 frames per commercial'),('Research sources','=COUNTA(Research!B7:B10)','Use limits retained'),('Pending human gates','=COUNTIF(Decision_Gates!I7:I36,"Pending")','No external release authorised')]
for ix,(lab,formula,note) in enumerate(metrics):
    r=7+(ix//3)*4;c=2+(ix%3)*3;ov.merge_cells(start_row=r,start_column=c,end_row=r,end_column=c+1);ov.cell(r,c,lab).font=Font(name='Georgia',size=10,bold=True,color=BLACK);ov.cell(r,c).fill=PatternFill('solid',fgColor=ROSE);ov.merge_cells(start_row=r+1,start_column=c,end_row=r+1,end_column=c+1);ov.cell(r+1,c,formula).font=Font(name='Georgia',size=17,bold=True,color=PLUM);ov.merge_cells(start_row=r+2,start_column=c,end_row=r+2,end_column=c+1);ov.cell(r+2,c,note).font=Font(name='Calibri',size=8,italic=True,color=GREY)
for r,txt in enumerate(['• Update the same CMP/COM IDs and version fields; do not create a new tracker for revisions.','• Social/TikTok preparation requires client-approved product facts, claims, account access, rights, privacy, moderation, target/audience and measurement settings.','• No paid post, creator outreach, UGC reuse, product sampling, data capture, clinic/procedure, product claim or retail/commerce activity proceeds without named human gates.','• Replace every planning scenario and budget with client data, supplier quotations and approved media/creator plans before commercial sign-off.'],21):
    ov.merge_cells(start_row=r,start_column=2,end_row=r,end_column=9);ov.cell(r,2,txt).font=Font(name='Calibri',size=9,italic=True,color=GREY);ov.row_dimensions[r].height=24
for col,w in {'B':26,'C':16,'D':3,'E':26,'F':16,'G':3,'H':26,'I':16}.items():ov.column_dimensions[col].width=w
ws=wb.create_sheet('Campaigns');style(ws);title(ws,'Campaign / Activation Register','One row per campaign. Planning-only values require client replacement and approval.',22)
rows=[[c['id'],c['shortName'],c['titleMm'],c['format'],c['commercialTension'],c['audience'],c['conversionMechanism'],c['creativeTerritory'],c['seasonalLogic'],'Concept ready',c['budgetMMK']['lean'],c['budgetMMK']['recommended'],c['budgetMMK']['flagship'],'Client-defined','; '.join(c['primaryKpis']),' ; '.join(c['requiredApprovals']),', '.join(c['sourceUse']),'Drive upload pending','v1.0','Planning only'] for c in CAM]
s,e=grid(ws,6,['Campaign ID','Name','Myanmar title','Format','Commercial tension','Audience / behaviour change','Conversion mechanism','Creative territory','Seasonal logic','Status','Lean MMK','Recommended MMK','Flagship MMK','Primary conversion','KPIs','Approvals','Source IDs','Latest artifact','Current version','Notes'],rows,[19,26,30,30,35,34,38,28,28,16,14,16,14,20,35,42,16,25,14,18]);ws.auto_filter.ref=f'B6:U{e}';ws.freeze_panes='B7'
for col in ['L','M','N']:
    for r in range(s,e+1):ws[f'{col}{r}'].number_format='MMK #,##0'
ws=wb.create_sheet('Social_TikTok');style(ws);title(ws,'Social Media & TikTok Preparation Register','A preparation workstream within each campaign. It must be approved before any organic/paid/creator/UGC/action.',17)
rows=[[c['id'],c['shortName'],c['socialMediaPreparation']['role'],c['socialMediaPreparation']['organicPreparation'],c['socialMediaPreparation']['tiktokPreparation'],c['socialMediaPreparation']['measurementPreparation'],'Not started','G1 facts/claims + G4 rights/privacy/policy','No publication or paid distribution authorised','Client social / legal / data owner','v1.0','Drive upload pending','Planning only'] for c in CAM]
s,e=grid(ws,6,['Campaign ID','Campaign','Channel role','Organic preparation','TikTok-specific preparation','Measurement preparation','Status','Required gate','Control statement','Owner','Version','Latest artifact','Notes'],rows,[19,25,26,44,48,38,16,28,34,25,14,25,22]);ws.auto_filter.ref=f'B6:N{e}';ws.freeze_panes='B7'
ws=wb.create_sheet('Commercials');style(ws);title(ws,'Standalone Commercial & Storyboard Register','COM records remain separate from linked campaign, including production, social versioning and rights decisions.',18)
rows=[[c['id'],c['titleEn'],c['titleMm'],c['format'],c['linkedCampaign'],c['tension'],c['objective'],c['territory'],c['hook'],'Client-approved CTA required',c['storyboardStatus'],len(c['storyboard']),'Concept storyboard ready','G5 product facts, claims, cast, locations, rights, music, subtitles, social versions, safety and privacy','Drive upload pending','v1.0'] for c in COM]
s,e=grid(ws,6,['Commercial ID','Title','Myanmar title','Format','Linked campaign','Tension','Objective','Creative territory','Hook','CTA','Storyboard status','Frames','Status','Production gate','Latest artifact','Current version'],rows,[19,28,29,25,20,34,34,30,34,30,32,12,22,52,25,14]);ws.auto_filter.ref=f'B6:Q{e}';ws.freeze_panes='B7'
ws=wb.create_sheet('Research');style(ws);title(ws,'Research & Evidence Register','Public sources are context only. Read use limits before applying any source to client copy.',10)
research=[('BEA-S01','TikTok For Business','Official beauty guide','Beauty awareness, interaction, commerce/lead routes and organic-support context','No proof of client account, Shop/feature access, eligibility, performance, cost, conversion or local availability','https://ads.tiktok.com/business/en-US/blog/beauty-advertising-tiktok-guide','21 Aug 2026','Platform context','Verified'),('BEA-S02','Food and Drug Administration, Myanmar','Official agency site','Agency states a safety/quality role for cosmetics and identifies a cosmetic control division','No proof any client product, label, claim, sample, activity or campaign is approved','https://www.fda.gov.mm/','21 Aug 2026','Regulatory boundary','Verified'),('BEA-S03','TikTok For Business','Official advertising policy','Healthcare/cosmetic-procedure policy boundary and restrictions context','Exact eligibility is market/client/account/creative specific; do not infer approval','https://ads.tiktok.com/resources/help/article/tiktok-ads-policy-healthcare-pharmaceuticals','21 Aug 2026','Policy boundary','Verified'),('BEA-S04','Lebo Lion','Third-party YouTube masterclass','Contextual opinions on profile, calendar, hook, CTA, engagement and analytics workflow','Non-Myanmar third-party opinion; not platform policy, local proof or client performance evidence','https://www.youtube.com/watch?v=Wa7aHXpMw_4','21 Aug 2026','Workflow context','Context only')]
grid(ws,6,['Source ID','Source','Type','Finding retained','Use limitation','URL','Date','Use','Status'],research,[12,28,22,38,46,52,14,20,16]);ws.auto_filter.ref='B6:J10';ws.freeze_panes='B7'
for r in range(7,11):ws.cell(r,7).hyperlink=ws.cell(r,7).value;ws.cell(r,7).font=Font(color='0563C1',underline='single',size=9)
ws=wb.create_sheet('Decision_Gates');style(ws);title(ws,'Human Decision Gates','Named human owners must clear each gate before external action.',10);rows=[]
for c in CAM:
    n=c['id'].split('-')[-1];rows += [[f'GATE-{n}-G1',c['id'],'Campaign','G1 Facts & claims','Are product/service, claims, label/pack, audience and evidence boundaries approved?','Client brand/regulatory owner','Before content/activation','Pending','Approved factual brief'],[f'GATE-{n}-G4',c['id'],'Campaign','G4 Social/activation feasibility','Are venue, hygiene, creators, rights, privacy, moderation, platform/policy, data and escalation owners named?','Client operations/social/legal owner','Before publication/activation','Pending','Approved preflight register']]
for c in COM:
    n=c['id'].split('-')[-1];rows.append([f'GATE-{n}-G5',c['id'],'Commercial','G5 Production go/no-go','Are script, product facts, claims, cast, location, rights, pack, music, social versions, safety, subtitles and CTA approved?','Client brand/legal owner','Before shoot','Pending','Approved production pack'])
grid(ws,6,['Gate record','Record ID','Track','Gate','Decision question','Owner','Due / timing','Status','Evidence required'],rows,[18,20,16,25,50,28,20,16,42]);ws.auto_filter.ref='B6:J36';ws.freeze_panes='B7';dv=DataValidation(type='list',formula1='"Pending,Approved,Hold,Rejected,Completed"');ws.add_data_validation(dv);dv.add('I7:I36')
ws=wb.create_sheet('Portfolio_Input');style(ws);title(ws,'Live Tracker Import Rows','Canonical-ID reference for the Master Tracker; use the same records and version fields after human approval.',14)
rows=[]
for c in CAM:rows.append([c['id'],'Campaign',c['shortName'],c['format'],'G2 Strategy Selected','Concept ready',c['budgetMMK']['recommended'],c['conversionMechanism'],c['socialMediaPreparation']['role'],'v1.0','ZYNTH-20260821-BEAUTY-BILINGUAL','Planning/public context'])
for c in COM:rows.append([c['id'],'Commercial',c['titleEn'],c['format'],'G3 Creative Selected','Concept storyboard ready','n/a','Client-approved CTA required','12-frame storyboard; social versions gated','v1.0','ZYNTH-20260821-BEAUTY-BILINGUAL','Planning/public context'])
grid(ws,6,['Record ID','Track','Title','Format','Gate','Status','Recommended budget/n.a.','Primary conversion/CTA','Social/TikTok note','Version','Batch code','Classification'],rows,[20,15,30,30,20,23,20,44,38,14,32,26]);ws.auto_filter.ref='B6:M26';ws.freeze_panes='B7'
chart=BarChart();chart.type='bar';chart.style=10;chart.title='Recommended Campaign Planning Envelope (MMK)';chart.y_axis.title='Campaign';chart.x_axis.title='MMK';data=Reference(wb['Campaigns'],min_col=13,max_col=13,min_row=6,max_row=16);cats=Reference(wb['Campaigns'],min_col=3,max_col=3,min_row=7,max_row=16);chart.add_data(data,titles_from_data=True);chart.set_categories(cats);chart.height=8;chart.width=18;chart.series[0].graphicalProperties.solidFill=GOLD;ov.add_chart(chart,'B27')
for ws in wb.worksheets:style(ws)
out=MON/'ZYNTH-20260821-BEAUTY-Monitoring.xlsx';wb.save(out)
rec=sum(c['budgetMMK']['recommended'] for c in CAM);report=['# ZYNTH Beauty / Cosmetics Monitoring Report','','**Batch:** `ZYNTH-20260821-BEAUTY-BILINGUAL`','','## Portfolio Monitor','','- Campaign proposals: **10**','- Mandatory campaign social/TikTok preparation routes: **10**','- Standalone commercial/storyboard concepts: **10**','- Detailed storyboard frames: **120**','- Research-source records: **4**','',f'- Recommended aggregate planning envelope: **{mmk(rec)}**','', '> Planning envelopes are non-tax assumptions, not media plans, quotations, forecasts or guarantees.','', '## Controlled Use Rules','', '- Social media and TikTok preparation is part of each campaign; it cannot replace the separate COM storyboard/production gate.','- No efficacy, safety, medical, clinical, before/after, testimonial, product, creator, retail, price, availability or platform-result claim is authorised without written client approval.','- No paid media, creator outreach, UGC reuse, data capture, sampling, platform targeting, shoot or activation proceeds without named human owners and gate clearance.']
(MON/'monitoring_report.md').write_text('\n'.join(report)+'\n',encoding='utf-8')
files=[]
for p in sorted(ROOT.rglob('*')):
    if p.is_file() and p.name!='source_manifest.json':files.append({'path':str(p.relative_to(ROOT)),'bytes':p.stat().st_size,'sha256':hashlib.sha256(p.read_bytes()).hexdigest()})
(MON/'source_manifest.json').write_text(json.dumps({'batchCode':'ZYNTH-20260821-BEAUTY-BILINGUAL','counts':{'campaignProposals':10,'socialTikTokPreparationRoutes':10,'commercialConcepts':10,'storyboardFrames':120,'wordDocuments':20,'markdownDocuments':21,'researchSources':4},'files':files},ensure_ascii=False,indent=2)+'\n',encoding='utf-8')
print(out);print('recommended portfolio envelope',rec)

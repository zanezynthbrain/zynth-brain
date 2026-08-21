from pathlib import Path
import json, re, hashlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import DataBarRule
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path('/home/ubuntu/zynth-brain/backend/outputs/zynth_batches/2026-08-21-fnb-bilingual')
MON = ROOT/'monitoring'; MON.mkdir(exist_ok=True)
campaigns = json.loads((ROOT/'data/campaigns.json').read_text(encoding='utf-8'))['campaigns']
commercials = json.loads((ROOT/'data/commercials.json').read_text(encoding='utf-8'))['commercials']

PRIMARY='5D4037'; LIGHT='E6DDD9'; ACCENT='A75E2B'; PALE='FFF8F2'; TEAL='1A6A68'; GREY='666666'; WHITE='FFFFFF'
header_fill=PatternFill('solid',fgColor=PRIMARY); light_fill=PatternFill('solid',fgColor=LIGHT); pale_fill=PatternFill('solid',fgColor=PALE)
header_font=Font(name='Calibri',size=10,bold=True,color=WHITE); title_font=Font(name='Georgia',size=20,bold=True,color=PRIMARY); section_font=Font(name='Georgia',size=13,bold=True,color=PRIMARY); normal_font=Font(name='Calibri',size=10,color='1F2933'); note_font=Font(name='Calibri',size=9,italic=True,color=GREY)
thin=Side(style='thin',color='D6D6D6'); med=Side(style='medium',color=PRIMARY)

def style_sheet(ws):
    ws.sheet_view.showGridLines=False
    ws.column_dimensions['A'].width=3
    for row in ws.iter_rows():
        for c in row:
            c.font=normal_font
            c.alignment=Alignment(vertical='center',wrap_text=True)

def add_title(ws,title,subtitle,last_col):
    ws.merge_cells(start_row=2,start_column=2,end_row=2,end_column=last_col)
    ws.cell(2,2,title).font=title_font; ws.cell(2,2).alignment=Alignment(vertical='center')
    ws.row_dimensions[2].height=30
    ws.merge_cells(start_row=3,start_column=2,end_row=3,end_column=last_col)
    ws.cell(3,2,subtitle).font=note_font; ws.cell(3,2).alignment=Alignment(wrap_text=True,vertical='center')
    ws.row_dimensions[3].height=28

def make_table(ws,row,headers,values,widths):
    for j,h in enumerate(headers,2):
        c=ws.cell(row,j,h); c.fill=header_fill; c.font=header_font; c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); c.border=Border(top=med,bottom=med,left=thin,right=thin)
    ws.row_dimensions[row].height=28
    for i,record in enumerate(values,row+1):
        for j,val in enumerate(record,2):
            c=ws.cell(i,j,val); c.font=normal_font; c.alignment=Alignment(vertical='center',wrap_text=True)
            c.fill=pale_fill if (i-row)%2 else PatternFill('solid',fgColor='FFFFFF')
            c.border=Border(top=thin,bottom=thin,left=thin,right=thin)
        ws.row_dimensions[i].height=44
    for j,w in enumerate(widths,2): ws.column_dimensions[chr(64+j)].width=w
    return row+1,row+len(values)

def parse_number(s):
    m=re.search(r'(\d+)',s or '')
    return int(m.group(1)) if m else None

wb=Workbook(); ws=wb.active; ws.title='Overview'; style_sheet(ws)
add_title(ws,'ZYNTH F&B Two-Hour Batch — Command View','Batch: ZYNTH-20260821-FNB-BILINGUAL | Created 2026-08-21 UTC | All amounts/scenarios are non-tax planning assumptions, not quotes or guarantees.',9)
ws['B5']='LIVE PORTFOLIO SNAPSHOT'; ws['B5'].font=section_font
metrics=[('Campaign proposals', '=COUNTA(Campaigns!B7:B16)','Ten client-ready campaign / activation concepts'),('Commercial concepts', '=COUNTA(Commercials!B7:B16)','Ten separate commercial ideas'),('Storyboard frames', '=SUM(Commercials!L7:L16)','Twelve detailed frames per concept'),('Research sources logged', '=COUNTA(Research!B7:B16)','Public/context sources; use limitations recorded'),('Recommended portfolio budget', '=SUM(Campaigns!N7:N16)','MMK; non-tax planning envelope'),('Pending human approvals', '=COUNTIF(Decision_Gates!H7:H36,"Pending")','Required before production or external action')]
for idx,(label,formula,note) in enumerate(metrics):
    r=7+(idx//3)*4; col=2+(idx%3)*3
    ws.merge_cells(start_row=r,start_column=col,end_row=r,end_column=col+1)
    ws.cell(r,col,label).font=Font(name='Georgia',size=10,bold=True,color=PRIMARY); ws.cell(r,col).fill=light_fill
    ws.merge_cells(start_row=r+1,start_column=col,end_row=r+1,end_column=col+1)
    ws.cell(r+1,col,formula).font=Font(name='Georgia',size=18,bold=True,color=TEAL); ws.cell(r+1,col).fill=PatternFill('solid',fgColor='FFFFFF')
    if 'budget' in label.lower(): ws.cell(r+1,col).number_format='MMK #,##0'
    ws.merge_cells(start_row=r+2,start_column=col,end_row=r+2,end_column=col+1)
    ws.cell(r+2,col,note).font=note_font; ws.cell(r+2,col).alignment=Alignment(wrap_text=True,vertical='center')
    for rr in range(r,r+3):
        for cc in range(col,col+2): ws.cell(rr,cc).border=Border(top=thin,bottom=thin,left=thin,right=thin)
        ws.row_dimensions[rr].height=22
ws['B20']='KEY USE RULES'; ws['B20'].font=section_font
notes=['• Keep the same CMP/COM IDs and update their current version rather than starting a new workbook.','• Scenario values require client funnel, margin, capacity, inventory and attribution data before commercial approval.','• No client, venue, chef, sponsor, product claim, price, availability, safety, food label, permit, partnership, right or legal status is confirmed by this batch.','• Use the G0–G7 gate register for human decisions; no external activity proceeds without the named owner.']
for i,n in enumerate(notes,21): ws.merge_cells(start_row=i,start_column=2,end_row=i,end_column=9); ws.cell(i,2,n).font=note_font; ws.cell(i,2).alignment=Alignment(wrap_text=True,vertical='center'); ws.row_dimensions[i].height=24
for col,w in {'B':25,'C':16,'D':3,'E':25,'F':16,'G':3,'H':25,'I':16}.items(): ws.column_dimensions[col].width=w

# Campaigns
ws=wb.create_sheet('Campaigns'); style_sheet(ws); add_title(ws,'Campaign / Activation Register','One row per campaign proposal. Update only this canonical row after human review.',21)
headers=['Campaign ID','Name','Myanmar title','Format','Commercial tension','Objective','Creative territory','Conversion mechanism','Seasonal logic','Status','Lean MMK','Recommended MMK','Flagship MMK','Base qualified follow-ups','Primary KPI','Approvals','Source IDs','Latest artifact','Current version','Notes']
rows=[]
for c in campaigns:
    base=parse_number(c['planningOutcomes']['base'].split('/')[-1])
    rows.append([c['id'],c['shortName'],c['titleMm'],c['format'],c['commercialTension'],c['objective'],c['creativeTerritory'],c['conversionMechanism'],c['seasonalLogic'],'Concept ready',c['budgetMMK']['lean'],c['budgetMMK']['recommended'],c['budgetMMK']['flagship'],base,'; '.join(c['primaryKpis']),'; '.join(c['requiredApprovals']),', '.join(c['sourceUse']),'Drive upload pending','v1.0','Planning only'])
start,end=make_table(ws,6,headers,rows,[18,24,28,26,35,30,28,35,28,16,14,16,14,18,30,35,16,26,14,22])
for col in ['K','L','M']:
    for r in range(start,end+1): ws[f'{col}{r}'].number_format='MMK #,##0'
ws.auto_filter.ref=f'B6:U{end}'; ws.freeze_panes='B7'; ws.conditional_formatting.add(f'L{start}:L{end}',DataBarRule(start_type='min',end_type='max',color=ACCENT))

# Commercials
ws=wb.create_sheet('Commercials'); style_sheet(ws); add_title(ws,'Standalone Commercial & Storyboard Register','Commercials are independent COM records even where strategically linked to a campaign.',17)
headers=['Commercial ID','Title','Myanmar title','Format','Linked campaign','Tension','Objective','Creative territory','Hook','CTA','Storyboard status','Frames','Status','Production gate','Latest artifact','Current version']
rows=[]
for c in commercials:
    rows.append([c['id'],c['titleEn'],c['titleMm'],c['format'],c['linkedCampaign'],c['tension'],c['objective'],c['territory'],c['hook'],c['cta'],c['storyboardStatus'],len(c['storyboard']),'Concept storyboard ready','Client facts, legal/claims, food safety, cast, locations, rights, music, subtitles and end-card required','Drive upload pending','v1.0'])
start,end=make_table(ws,6,headers,rows,[18,26,28,28,20,34,30,30,34,30,35,12,20,42,25,14])
ws.auto_filter.ref=f'B6:Q{end}'; ws.freeze_panes='B7'; ws.conditional_formatting.add(f'L{start}:L{end}',DataBarRule(start_type='min',end_type='max',color=TEAL))

# Research
ws=wb.create_sheet('Research'); style_sheet(ws); add_title(ws,'Research & Evidence Register','Each source is public context only. Read the use limitation before applying it to any proposal.',9)
research=[
('FNB-S01','USDA FAS — Food Service', 'Government market-report landing page','Food-service sector scope/context','No detailed local claims retained','https://www.fas.usda.gov/data/gain/2025/02/burma-food-service-hotel-restaurant-institutional','2025-02-11','Context only','Verified'),
('FNB-S02','Xinhua — Tastes of Golden Land','Public media/event coverage','Traditional food show mechanics, regional products, MSME networking','Historical event only; no future partnership/venue claim','https://english.news.cn/20260726/49ff2f8a557b4825a58978e89ccccc61/c.html','2026-07-26','Event format context','Verified'),
('FNB-S04','Myanmar Government — Upcoming Holidays','Official calendar','Q4 2026 holiday guardrails','Calendar only; not demand/operations proof','https://myanmar.gov.mm/upcoming-holidays','2026-08-21','Calendar','Verified'),
('FNB-S05','Myanmar eVisa — Public Holidays','Official calendar','Q4 date corroboration','Calendar only','https://evisa.moip.gov.mm/notice/public-holiday','2026-08-21','Calendar','Verified'),
('FNB-S06','Tilleke — Myanmar-language labeling','Public legal commentary','Packaging/label review reminder','Not legal advice/current product determination','https://www.tilleke.com/insights/myanmar-language-labeling-required-wide-range-products/','2018-10-26','Compliance boundary','Verified context'),
('FNB-S07','MITV — Tastes of Golden Land','Official local coverage','Competition/exhibition/sales format context','Historical only','https://www.myanmaritv.com/news/tastes-golden-land-myanmar-traditional-food-competition-and-exhibition','2026-07-26','Event format context','Verified'),
('FNB-S08','MOI — Traditional Food Exhibition','Official publisher','Panels, associations, demos, booths context','No forecast/partnership claim','https://www.moi.gov.mm/moi%3Aeng/news/21664','2026-07-27','Event format context','Verified'),
('FNB-S09','GNLM — Signature Foods','Local publisher/MNA coverage','Fine dining/regional-display/knowledge-exchange context','No future sponsor/endorsement claim','https://www.gnlm.com.mm/myanmar-urged-to-cultivate-signature-foods/','2026-07-26','Cultural/format context','Verified'),
('FNB-S10','Inter Myanmar Channel — HORECA video','First-hand public video','Observed demo, skills, tasting, seating/stage mechanics','All named event/brand/date/sponsor claims separately unverified','https://www.youtube.com/watch?v=mbh1472c2HI','2026-08-21','Observed mechanics only','Use with limitation')]
make_table(ws,6,['Source ID','Source','Type','Finding retained','Use limitation','URL','Date','Use','Status'],research,[12,28,22,34,42,52,14,22,18]); ws.auto_filter.ref='B6:J15'; ws.freeze_panes='B7'
for r in range(7,16): ws.cell(r,7).hyperlink=ws.cell(r,7).value; ws.cell(r,7).font=Font(color='0563C1',underline='single',size=9)

# Gate register
ws=wb.create_sheet('Decision_Gates'); style_sheet(ws); add_title(ws,'Human Decision Gates','This is the non-automated control layer. Named human owners must clear each gate before the next production action.',10)
headers=['Gate record','Record ID','Track','Gate','Decision question','Owner','Due / timing','Status','Evidence required']
values=[]
for c in campaigns:
    values.append([f'GATE-{c["id"].split("-")[-1]}-G1',c['id'],'Campaign','G1 Research Cleared','Are public facts, use limitations and product/claims boundaries sufficient for concept development?','Client evidence owner','Before development','Pending','Verified source review + factual product brief'])
    values.append([f'GATE-{c["id"].split("-")[-1]}-G4',c['id'],'Campaign','G4 Feasibility','Are food safety, allergy, hygiene, venue, consent, rights, price/inventory and permit owners named?','Client operations owner','Before production','Pending','Approved preflight register'])
for c in commercials:
    values.append([f'GATE-{c["id"].split("-")[-1]}-G5',c['id'],'Commercial','G5 Go / No-go','Are script, claims, cast, location, food styling/handling, music, rights, subtitles and CTA approved?','Client brand/legal owner','Before shoot','Pending','Approved script and production pack'])
make_table(ws,6,headers,values,[18,18,16,16,42,24,18,16,38]); ws.auto_filter.ref='B6:J36'; ws.freeze_panes='B7'
dv=DataValidation(type='list',formula1='"Pending,Approved,Hold,Rejected,Completed"',allow_blank=False); ws.add_data_validation(dv); dv.add('I7:I36')

# Portfolio input
ws=wb.create_sheet('Portfolio_Input'); style_sheet(ws); add_title(ws,'Live Tracker Import Rows','Paste these values into the live ZYNTH Master Tracker only after a human reviews/approves each row.',14)
headers=['Record ID','Track','Title','Format','Gate','Status','Recommended budget / n.a.','Primary conversion or CTA','Source/asset note','Version','Batch code','Data classification']
values=[]
for c in campaigns: values.append([c['id'],'Campaign',c['shortName'],c['format'],'G2 Strategy Selected','Concept ready',c['budgetMMK']['recommended'],c['conversionMechanism'],', '.join(c['sourceUse']),'v1.0','ZYNTH-20260821-FNB-BILINGUAL','Planning / public context'])
for c in commercials: values.append([c['id'],'Commercial',c['titleEn'],c['format'],'G3 Creative Selected','Concept storyboard ready','n/a',c['cta'],'12-frame storyboard in batch package','v1.0','ZYNTH-20260821-FNB-BILINGUAL','Planning / public context'])
make_table(ws,6,headers,values,[20,15,28,28,20,22,20,42,28,14,30,26]); ws.auto_filter.ref='B6:M26'; ws.freeze_panes='B7'
for r in range(7,27):
    if isinstance(ws.cell(r,8).value,(int,float)): ws.cell(r,8).number_format='MMK #,##0'

# add portfolio chart to Overview after campaign data exists
ws=wb['Overview']; chart=BarChart(); chart.type='bar'; chart.style=10; chart.title='Recommended Campaign Planning Envelope (MMK)'; chart.y_axis.title='Campaign'; chart.x_axis.title='MMK';
data=Reference(wb['Campaigns'],min_col=13,max_col=13,min_row=6,max_row=16); cats=Reference(wb['Campaigns'],min_col=3,max_col=3,min_row=7,max_row=16); chart.add_data(data,titles_from_data=True); chart.set_categories(cats); chart.height=8; chart.width=18; chart.series[0].graphicalProperties.solidFill=ACCENT; ws.add_chart(chart,'B27')

for sheet in wb.worksheets:
    style_sheet(sheet)
    for row in sheet.iter_rows():
        for c in row:
            if c.value is not None and c.row>=6: c.alignment=Alignment(vertical='center',wrap_text=True)

out=MON/'ZYNTH-20260821-FNB-Monitoring.xlsx'; wb.save(out)

# Markdown monitoring report
campaign_total=sum(c['budgetMMK']['recommended'] for c in campaigns)
lean_total=sum(c['budgetMMK']['lean'] for c in campaigns); flag_total=sum(c['budgetMMK']['flagship'] for c in campaigns)
md=['# ZYNTH F&B Monitoring Report','',f'**Batch:** `ZYNTH-20260821-FNB-BILINGUAL`  ',f'**Generated:** 2026-08-21 UTC  ','', '## Portfolio Monitor','',f'- Campaign proposals: **{len(campaigns)}**','- Standalone commercial/storyboard concepts: **10**','- Detailed storyboard frames: **120**','- Research-source records: **9**','', '## Portfolio Planning Envelopes','',f'- Lean: **MMK {lean_total/1e6:.1f}m**','- Recommended: **MMK {campaign_total/1e6:.1f}m**','- Flagship: **MMK {flag_total/1e6:.1f}m**','', '> These are non-tax planning envelopes, not supplier quotations. Financial, attendance, lead and conversion outcomes are directional assumptions only; replace them with client data before sign-off.','', '## Monitoring Coverage','', 'The workbook tracks campaign and commercial IDs separately; tension, conversion/CTA, creative territory, planning budget, scenarios, required approvals, research-source limits, human gates, latest asset link and version.','', '## Non-Negotiable Controls','', '- Do not use public-event coverage as a partnership, sponsor, venue, turnout, sales or endorsement claim.','- Route product facts, ingredients, allergens, labels, food safety, prices, availability, rights, permits, consent and legal terms to named client owners.','- The AI Council can provide structured ideas and critique, but human decision records remain authoritative.']
(MON/'monitoring_report.md').write_text('\n'.join(md)+'\n',encoding='utf-8')

# hash manifest
files=[]
for p in sorted(ROOT.rglob('*')):
    if p.is_file() and '.git' not in p.parts:
        files.append({'path':str(p.relative_to(ROOT)),'bytes':p.stat().st_size,'sha256':hashlib.sha256(p.read_bytes()).hexdigest()})
manifest={'batchCode':'ZYNTH-20260821-FNB-BILINGUAL','generatedAt':'2026-08-21T06:45:00Z','counts':{'campaignProposals':10,'commercialConcepts':10,'storyboardFrames':120,'wordDocuments':20,'markdownDocuments':21,'researchSources':9},'files':files}
(MON/'source_manifest.json').write_text(json.dumps(manifest,indent=2)+'\n',encoding='utf-8')
print(out)
print('Recommended portfolio planning envelope:', campaign_total)
